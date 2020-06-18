# -*- coding: utf-8 -*-
from __future__ import unicode_literals

import os, re, random, string,xlrd,sys,xlwt
from openpyxl import Workbook
import datetime
from decouple import config
from django.contrib.auth.models import User
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.core.mail import EmailMultiAlternatives
from django.forms import ValidationError
from django.http import JsonResponse,HttpResponse
from django.shortcuts import render,redirect, reverse
from django.template.loader import render_to_string
from django.utils.timezone import now as timezone_now
from django.utils.encoding import smart_str, smart_unicode

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import authentication, permissions

from .guards import *

from arbitrajes.models import Arbitro
from arbitrajes.views import list_trabajos_aceptados
from autores.forms import ImportFromExcelForm
from autores.models import Autores_trabajos, Autor
from autores.views import get_extension_file, validate_alpha
from main_app.forms import UploadFileForm
from main_app.models import Usuario_rol_in_sistema, Sistema_asovac, Area,Sub_area
from main_app.views import (
    get_route_resultados, get_route_trabajos_navbar, get_route_trabajos_sidebar,
    get_roles, get_route_configuracion, get_route_seguimiento,
    validate_rol_status, handle_uploaded_file
)
from main_app.decorators import user_is_arbitraje
from trabajos.models import Trabajo_arbitro

from .utils import (
    CertificateGenerator, MiscellaneousGenerator, validate_recipients_excel
)
from .forms import (
    CertificateToRefereeForm, CertificateToAuthorsForm, MultipleRecipientsForm,
    MultipleRecipientsWithDateForm, MultipleRecipientsWithDateAndSubjectForm,
    MultipleRecipientsWithRoleForm,UploadFileForm,resumenContentForm,resumenPalabrasForm,UploadAficheForm,comisionAcademicaForm
)

from .models import Resumen
from eventos.models import Evento,Locacion_evento,Organizador
from reportlab.pdfgen import canvas
from io import BytesIO

from reportlab.platypus import SimpleDocTemplate, Paragraph, TableStyle, Image,Spacer,Table
from reportlab.platypus import BaseDocTemplate, Frame, NextPageTemplate, PageBreak, PageTemplate
from reportlab.lib.styles import getSampleStyleSheet,ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.enums import TA_LEFT,TA_RIGHT,TA_CENTER,TA_JUSTIFY

# Importamos clase de hoja de estilo de ejemplo.
# Se importa el tamaño de la hoja.
from reportlab.lib.pagesizes import A3
from reportlab.lib.pagesizes import A4
from datetime import datetime
from functools import partial


MONTH_NAMES = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio',
               'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']

DAY_NAMES= ['Lunes','Martes','Miércoles','Jueves','Viernes','Sábado','Domingo']

TOP_NAVBAR_OPTIONS = [{'title':'Configuración', 'icon': 'fa-cogs', 'active': True},
                      {'title':'Monitoreo', 'icon': 'fa-eye', 'active': False},
                      {'title':'Resultados', 'icon': 'fa-chart-area','active': False},
                      {'title':'Administración', 'icon': 'fa-archive', 'active': False}]

class ChartData(APIView):
    authentication_classes = ()
    permission_classes = ()

    def get(self, request, mode, format=None):
        arbitraje_id = request.session['arbitraje_id']
        #El siguiente codigo es para obtener los datos necesarios para la grafica de resultados de trabajos
        trabajos_aceptados = Autores_trabajos.objects.filter(sistema_asovac = arbitraje_id, pagado = True, es_autor_principal = True, trabajo__estatus = "Aceptado").count()
        trabajos_rechazados = Autores_trabajos.objects.filter(sistema_asovac = arbitraje_id, pagado = True, es_autor_principal = True, trabajo__estatus = "Rechazado").count()
        trabajos_pendientes = Autores_trabajos.objects.filter(sistema_asovac = arbitraje_id, pagado = True, es_autor_principal = True, trabajo__estatus = "Pendiente").count()
        
        #El siguiente codigo es para obtener los datos necesarios para la grafica de datos de arbitrajes y datos de area
        total_arbitros = Usuario_rol_in_sistema.objects.filter(rol = 4, sistema_asovac = arbitraje_id, status = True).count()
        autores_trabajos = Autores_trabajos.objects.filter(sistema_asovac = arbitraje_id, pagado = True)
        invitaciones_aceptadas = 0
        invitaciones_pendientes = 0
        
        areas_info = []
        areas_list = Area.objects.all() 
        for item in areas_list:
            areas_info.append({
                "id": item.id,
                "nombre": item.nombre,
                "trabajos_aceptados": 0,
                "trabajos_rechazados": 0,
                "cantidad_autores":0,
                "autores_id": []
                })
        
        for autor_trabajo in autores_trabajos:
            if autor_trabajo.es_autor_principal:    
                invitaciones_aceptadas += Trabajo_arbitro.objects.filter(trabajo = autor_trabajo.trabajo, invitacion = True).count()
                invitaciones_pendientes += Trabajo_arbitro.objects.filter(trabajo = autor_trabajo.trabajo, invitacion = False, fin_arbitraje = False).count()
            for item in areas_info:
                if autor_trabajo.trabajo.subareas.first().area.id == item["id"]:
                    if not autor_trabajo.autor.id in item["autores_id"]:
                        item["autores_id"].append(autor_trabajo.autor.id)
                    
                    if autor_trabajo.trabajo.estatus == "Aceptado" and autor_trabajo.es_autor_principal:
                        item["trabajos_aceptados"] += 1
                    elif autor_trabajo.trabajo.estatus == "Rechazado" and autor_trabajo.es_autor_principal:
                        item["trabajos_rechazados"] += 1


        areas_labels = []
        areas_trabajos_aceptados = []
        areas_trabajos_rechazados = []
        areas_autores = []
        areas_colores = []
        areas_labels_splited = []
        for item in areas_info:
            
            areas_labels_splited.append(abreviate_if_neccesary(item["nombre"]))
            areas_trabajos_aceptados.append(item["trabajos_aceptados"])
            areas_trabajos_rechazados.append(item["trabajos_rechazados"])

            if(len(item["autores_id"]) > 0):
                areas_labels.append(item["nombre"])
                areas_autores.append(len(item["autores_id"]))
                areas_colores.append(generate_random_color())

        #El siguiente codigo es para obtener los datos necesarios para la grafica de niveles de intruccion de los autores
        usuarios_autor = Usuario_rol_in_sistema.objects.filter(rol = 5, sistema_asovac = arbitraje_id, status = True)
        educacion_primaria = 0
        educacion_secundaria = 0
        bachillerato = 0
        tecnico = 0
        universidad = 0
        postgrado = 0
        for usuario_autor in usuarios_autor:
            try:
                autor = Autor.objects.get(usuario = usuario_autor.usuario_asovac)
                if autor.nivel_instruccion == "Educación Básica Primaria":
                    educacion_primaria += 1
                elif autor.nivel_instruccion == "Educación Básica Secundaria":
                    educacion_secundaria += 1
                elif autor.nivel_instruccion == "Bachillerato/Educación Media":
                    bachillerato += 1
                elif autor.nivel_instruccion == "Educación Técnico/Profesional":
                    tecnico += 1
                elif autor.nivel_instruccion == "Universidad":
                    universidad += 1
                else:
                    postgrado += 1
            except: 
                pass

        
        # Para definir propiedades del documento de excel
        if(mode == '1'):
            response = HttpResponse(content_type='application/ms-excel')
            response['Content-Disposition'] = 'attachment; filename=convencion_asovac_resultados.xls'
            style_detail = xlwt.easyxf('align: wrap on, horiz center; border : bottom thin,right thin,top thin,left thin;')
            workbook = xlwt.Workbook()
            worksheet = workbook.add_sheet("Resultados")
            # Para agregar los titulos de cada columna
            row_num = 1
            col_num = 1
            columns = ['Educación Básica Primaria', 'Educación Básica Secundaria', 'Bachillerato / Educación Media', 'Educación Técnico / Profesional', 'Universidad', 'Postgrado']
            columns_data = [educacion_primaria, educacion_secundaria, bachillerato, tecnico, universidad, postgrado]


            #Tabla para autores según nivel de instrucción
            worksheet.write_merge(0,0,0,len(columns), 'Autores según nivel de instrucción' , style_detail)
            worksheet.write(row_num, 0, '', style_detail)
            worksheet.write(row_num + 1, 0, 'Autores', style_detail)
            for column_number in range(len(columns)):
                worksheet.write(row_num, col_num, columns[column_number], style_detail)
                worksheet.write(row_num+1, col_num, columns_data[column_number], style_detail)
                col_num += 1
            

            #Tabla para resultados de trabajos
            col_num += 2
            second_col = col_num
            columns = ["Aceptados", "Rechazados", "Pendientes"]
            columns_data = [trabajos_aceptados, trabajos_rechazados, trabajos_pendientes]

            worksheet.write_merge(0,0,col_num, col_num + len(columns), 'Resultados de trabajos' , style_detail)
            worksheet.write(row_num, col_num, '', style_detail)
            worksheet.write(row_num + 1, col_num, 'Trabajos', style_detail)
            col_num += 1

            for column_number in range(len(columns)):
                worksheet.write(row_num, col_num, columns[column_number], style_detail)
                worksheet.write(row_num + 1, col_num, columns_data[column_number], style_detail)
                col_num+= 1

            #Tabla para resultados de arbitrajes
            row_num += 4
            col_num = 0
            columns = ['Total árbitros', 'Invitaciones aceptadas', 'Invitaciones pendientes']
            columns_data = [total_arbitros, invitaciones_aceptadas, invitaciones_pendientes]

            worksheet.write_merge(row_num,row_num, col_num, len(columns), 'Resultados de arbitrajes', style_detail)
            worksheet.write(row_num + 1, col_num, '', style_detail)
            worksheet.write(row_num + 2, col_num, 'Cantidad', style_detail)
            col_num += 1

            for column_number  in range(len(columns)):
                worksheet.write(row_num + 1, col_num, columns[column_number], style_detail)
                worksheet.write(row_num + 2, col_num, columns_data[column_number], style_detail)
                col_num += 1

            #Tabla para resultados según áreas
            col_num = second_col
            columns = areas_labels_splited
            columns_aceptados = areas_trabajos_aceptados
            columns_rechazados = areas_trabajos_rechazados

            worksheet.write_merge(row_num, row_num, col_num, col_num + 2, 'Resultados por área', style_detail)

            row_num += 1
            worksheet.write(row_num, col_num, '', style_detail)
            worksheet.write(row_num, col_num + 1, 'Aprobados', style_detail)
            worksheet.write(row_num, col_num + 2, 'Rechazados', style_detail)

            row_num += 1
            for column_number in range(len(columns)):
                worksheet.write(row_num, col_num, columns[column_number], style_detail)
                worksheet.write(row_num, col_num + 1, columns_aceptados[column_number], style_detail)
                worksheet.write(row_num, col_num + 2, columns_rechazados[column_number], style_detail)
                row_num += 1
            

            #Tabla para autores por área

            row_num += 1
            col_num = 0

            columns = areas_labels
            columns_data = areas_autores

            worksheet.write_merge(row_num, row_num, col_num, len(columns), 'Autores por área', style_detail)

            row_num += 1
            worksheet.write(row_num, col_num, '', style_detail)
            worksheet.write(row_num + 1, col_num, 'Cantidad', style_detail)

            col_num += 1
            for column_number in range(len(columns)):
                worksheet.write(row_num, col_num, columns[column_number], style_detail)
                worksheet.write(row_num + 1, col_num, columns_data[column_number], style_detail)
                col_num += 1


            workbook.save(response)
            return response

        else:
            data = {
                "trabajos": {
                    "labels": ["Trabajos Aceptados", "Trabajos Rechazados", "Trabajos Pendientes"],
                    "data": [trabajos_aceptados, trabajos_rechazados, trabajos_pendientes]
                },
                "arbitros": {
                    "labels": ['Total árbitros', 'Invitaciones aceptadas', 'Invitaciones pendientes'],
                    "data": [total_arbitros, invitaciones_aceptadas, invitaciones_pendientes]
                },
                "autores": {
                    "labels": ['Educación Básica Primaria', 'Educación Básica Secundaria', 'Bachillerato/Educación Media', 'Educación Técnico/Profesional', 'Universidad', 'Postgrado'],
                    "data": [educacion_primaria, educacion_secundaria, bachillerato, tecnico, universidad, postgrado]
                },
                "area":{
                    "labels": areas_labels_splited,
                    "data_aceptados": areas_trabajos_aceptados,
                    "data_rechazados": areas_trabajos_rechazados,
                },
                "autoresArea":{
                    "labels": areas_labels,
                    "data": areas_autores,
                    "colores": areas_colores
                },
            }   
            return Response(data)


# Funcion de ayuda que genera un contexto comun para todas las vistas/views
def create_common_context(request):
    arbitraje_id = request.session['arbitraje_id']
    arbitraje = Sistema_asovac.objects.get(pk=arbitraje_id)
    estado = arbitraje.estado_arbitraje
    rol_id = get_roles(request.user.id, arbitraje_id)
    item_active = 1
    items = validate_rol_status(estado, rol_id, item_active, arbitraje_id)
    route_conf = get_route_configuracion(estado, rol_id, arbitraje_id)
    route_seg = get_route_seguimiento(estado, rol_id)
    route_trabajos_sidebar = get_route_trabajos_sidebar(estado, rol_id, item_active)
    route_trabajos_navbar = get_route_trabajos_navbar(estado, rol_id)
    route_resultados = get_route_resultados(estado, rol_id, arbitraje_id)

    context = {
        'main_navbar_options' : TOP_NAVBAR_OPTIONS,
        'estado' : estado,
        'rol_id' : rol_id,
        'arbitraje_id' : arbitraje_id,
        'item_active' : item_active,
        'items':items,
        'route_conf':route_conf,
        'route_seg':route_seg,
        'route_trabajos_sidebar':route_trabajos_sidebar,
        'route_trabajos_navbar': route_trabajos_navbar,
        'route_resultados': route_resultados,
    }

    return context


def generate_random_color():
    letras = "0123456789ABCDEF"
    color = "#"
    for x in range(0,6):
        color += letras[random.randint(0,15)]
    return color


def create_certificate_context(arbitraje):
    now = timezone_now()
    now_date_string = '%s de %s de %s' % (now.day, MONTH_NAMES[now.month - 1], now.year)
    path1 = arbitraje.firma1.url.replace('/media/', '')
    path2 = arbitraje.firma2.url.replace('/media/', '')
    path_logo = arbitraje.logo.url.replace('/media/', '')
    date_start_string = '%s de %s' % (arbitraje.fecha_inicio_arbitraje.day,
                                      MONTH_NAMES[arbitraje.fecha_inicio_arbitraje.month - 1])
    date_end_string = '%s de %s de %s' % (arbitraje.fecha_fin_arbitraje.day,
                                          MONTH_NAMES[arbitraje.fecha_fin_arbitraje.month - 1],
                                          arbitraje.fecha_fin_arbitraje.year)
    university_names = arbitraje.sedes.split(',')
    university_names = map(lambda n: n.strip(), university_names)
    # print university_names
    context = {}
    context["city"] = arbitraje.ciudad
    context["header_url"] = arbitraje.cabecera
    context["authority1_info"] = arbitraje.autoridad1.replace(',', '<br/>')
    context["signature1_path"] = os.path.join(os.path.join(settings.MEDIA_ROOT), path1)
    context["authority2_info"] = arbitraje.autoridad2.replace(',', '<br/>')
    context["signature2_path"] = os.path.join(os.path.join(settings.MEDIA_ROOT), path2)
    context["logo_path"] = os.path.join(os.path.join(settings.MEDIA_ROOT), path_logo)
    context['date_string'] = now_date_string
    context["roman_number"] = arbitraje.numero_romano
    context["date_start_string"] = date_start_string
    context["date_end_string"] = date_end_string
    context["university_names"] = university_names

    return context


def abreviate_if_neccesary(area_nombre):
    nombre_abreviado = ""
    if len(area_nombre) > 20:
        for item in area_nombre.split(" "):
            if item[0].isupper():
                nombre_abreviado += item[0]
        return nombre_abreviado
    return area_nombre


# Create your views here.
@login_required
@user_is_arbitraje
def recursos_pag(request):
    arbitraje_id = request.session['arbitraje_id']
    arbitraje = Sistema_asovac.objects.get(pk=arbitraje_id)
    estado = arbitraje.estado_arbitraje
    rol_id = get_roles(request.user.id, arbitraje_id)
    if not resultados_guard(estado, rol_id):
        raise PermissionDenied
    context = create_common_context(request)
    context['nombre_vista'] = 'Recursos'
    context['item_active'] = 3
    return render(request, 'recursos.html', context)


def get_data(request, *args, **kwargs):
    data = {
        "sales": 100,
        "customers": 10,
    }
    return JsonResponse(data)


@login_required
@user_is_arbitraje
def generate_pdf(request,*args , **kwargs):
    arbitraje_id = request.session['arbitraje_id']
    arbitraje = Sistema_asovac.objects.get(pk=arbitraje_id)
    start_date = arbitraje.fecha_inicio_arbitraje
    end_date = arbitraje.fecha_fin_arbitraje
    date_start_string = '%s de %s' % (start_date.day, MONTH_NAMES[start_date.month - 1])
    date_end_string = '%s de %s de %s' % (end_date.day, MONTH_NAMES[end_date.month - 1], end_date.year)

    filename = "AsoVAC_Rabindranath_Ferreira.pdf"

    context = {}
    context["date_start_string"] = date_start_string
    context["date_end_string"] = date_end_string
    context["university_names"] = ['Universidad Metropolitana', 'Universidad Central de Venezuela']
    context["header_url"] = arbitraje.cabecera
    context["city"] = 'Caracas'
    context["day"] = 4
    context["month"] = 'Agosto'
    context["year"] = 2018
    context["sex"] = 'F'
    context["full_name"] = 'Karla Calo'
    context["roman_number"] = 'LXVII'
    context["work_title"] = 'EVALUACIÓN DE LA OBTENCIÓN DE ACEITE ESENCIAL DE SARRAPIA (DIPTERYX ODORATA)\
     EMPLEANDO CO2 COMO FLUIDO SUPERCRÍTICO Y MACERACIÓN ASISTIDA CON ULTRASONIDO'
    context["work_code"] = 'BC-24'
    context["start_date"] = '29 de Noviembre'
    context["finish_date"] = '1 de Diciembre de 2018'
    context["convention_place"] = "Universidad Metropolintana (UNIMET) y la Universidad Central de Venezuela (UCV)"
    context["subject_title"] = 'Rabindranath Ferreira Villamizar'
    context["authors"] = ['Rabindranath Ferreira'] * 5
    #['Karla Calo', 'Luis Henríquez', 'Laura Margarita Febres', 'Yoleida Soto Anes','Karla Calo', 'Luis Henríquez', 'Laura Margarita Febres']
    context["max_info_date"] = '15 de Noviembre'
    context["observations"] =   ['El resumen no cumple con el formato indicado en la normativa de la Convención y el modelo enviado.',
                                'No se cumple con lo solicitado para el formato de nombres y apellidos de los autores completos, separados por coma.',
                                'La ventaja evidente del consumo de edulcorantes no calóricos es reducir el aporte calórico proveniente de sacarosa y \
                                otros endulzantes que sí aportan carbohidratos, pero ningún edulcorante conocido tiene poder hipoglicemiante. Tampoco tiene \
                                poder adelgazante per sé. Todo depende del resto del aporte de macronutrientes y otras condiciones de gasto calórico V/S requerimiento.',
                                'No están claros los porcentajes de ingesta definidos para los tres grupos.',
                                'Falta señalar cuantitativamente la reducción ponderal y en cuánto tiempo además de expresarlo con significancia y análisis estadístico.',
                                'Es conocido el after taste de la Stevia, sería interesante informar si los animales mostraron alguna preferencia o rechazo por las tres opciones propuestas.']
    context["deficient_areas"] = ['metodología', 'resultados', 'conclusiones', 'palabras clave']
    context["completed_work_form_link"] = 'https://docs.google.com/forms/d/e/1FAIpQLSdJsmghAty674AII18VKDbQDOv3-1b4jhZtJfqBouzYFwJL3g/viewform'
    context["footer_content"] = """http://www.asovac.org/lxvii-convencion-anual-de-asovac https://www.facebook.com/ConvencionAsovac2017<br/>(0212)753-5802 asovac.convencion2017@gmail.com"""
    doc_gen = MiscellaneousGenerator()

    return doc_gen.get_name_tag(filename, context)


@login_required
@user_is_arbitraje
def create_approval_letters(request):
    pass


@login_required
@user_is_arbitraje
def create_authors_certificates(request):
    arbitraje_id = request.session['arbitraje_id']
    arbitraje = Sistema_asovac.objects.get(pk=arbitraje_id)

    context = create_certificate_context(arbitraje)

    certificate_gen = CertificateGenerator()

    return certificate_gen.get_authors_certificate(context)


@login_required
@user_is_arbitraje
def resources_author(request):

    # print (rol_id)
    arbitraje_id = request.session['arbitraje_id']
    arbitraje = Sistema_asovac.objects.get(pk=arbitraje_id)
    estado = arbitraje.estado_arbitraje
    rol_id=get_roles(request.user.id , arbitraje_id)
    
    if not recursos_guard(estado, rol_id):
        raise PermissionDenied

    item_active = 1
    items=validate_rol_status(estado,rol_id,item_active, arbitraje_id)

    route_conf= get_route_configuracion(estado,rol_id, arbitraje_id)
    route_seg= get_route_seguimiento(estado,rol_id)
    route_trabajos_sidebar = get_route_trabajos_sidebar(estado,rol_id,item_active)
    route_trabajos_navbar = get_route_trabajos_navbar(estado,rol_id)
    route_resultados = get_route_resultados(estado,rol_id, arbitraje_id)
    # print items
    if request.method == "POST":
        form = CertificateToAuthorsForm(request.POST, sistema_id = arbitraje_id)
        if form.is_valid():
            cert_context = create_certificate_context(arbitraje)
            for trabajo_id in form.cleaned_data['trabajos']:
                autores_trabajo = Autores_trabajos.objects.filter(trabajo = trabajo_id)
                now = timezone_now()
                now_date_string = '%s de %s de %s' % (now.day, MONTH_NAMES[now.month - 1], now.year)
                authors_emails = []
                authors_names = []
                for autor_trabajo in autores_trabajo:
                    authors_emails.append(autor_trabajo.autor.correo_electronico)
                    authors_names.append( autor_trabajo.autor.nombres.split(' ')[0] + ' ' + autor_trabajo.autor.apellidos.split(' ')[0] )
                
                instance_context = {
                    'recipient_name': '',
                    'roman_number': arbitraje.numero_romano,
                    'subject_title': autores_trabajo.first().trabajo.titulo_espanol,
                    'people_names': authors_names
                }
                instance_context.update(cert_context)
                certificate_gen = CertificateGenerator()
                certificate = certificate_gen.get_authors_certificate(instance_context)

                filename = "Certificado_Convencion_Asovac_%s.pdf" % (instance_context["roman_number"])

                context_email = {
                    'sistema': arbitraje,
                    'titulo_trabajo': autores_trabajo.first().trabajo.titulo_espanol,
                    'rol': 'autor'
                }
                msg_plain = render_to_string('../templates/email_templates/certificate.txt', context_email)
                msg_html = render_to_string('../templates/email_templates/certificate.html', context_email)

                email_msg = EmailMultiAlternatives('Certificado de autor(es)', msg_plain, config('EMAIL_HOST_USER'), authors_emails)
                email_msg.attach(filename, certificate.content, 'application/pdf')
                email_msg.attach_alternative(msg_html, "text/html")
                email_msg.send()

            messages.success(request, 'Se han enviado los certificados al(los) autor(es) seleccionados con éxito.')
            return redirect('recursos:resources_author')
    else:    
        form = CertificateToAuthorsForm(sistema_id = arbitraje_id)
    context = {
        'nombre_vista' : 'Recursos',
        'main_navbar_options' : TOP_NAVBAR_OPTIONS,
        'estado' : estado,
        'rol_id' : rol_id,
        'arbitraje_id' : arbitraje_id,
        'item_active' : item_active,
        'items':items,
        'route_conf':route_conf,
        'route_seg':route_seg,
        'route_trabajos_sidebar':route_trabajos_sidebar,
        'route_trabajos_navbar': route_trabajos_navbar,
        'route_resultados': route_resultados,
        'form': form,
        'admin_styles': True
    }
    return render(request, 'recursos_author.html', context)


@login_required
@user_is_arbitraje
def resources_referee(request):

    arbitraje_id = request.session['arbitraje_id']
    arbitraje = Sistema_asovac.objects.get(pk=arbitraje_id)
    estado = arbitraje.estado_arbitraje
    rol_id=get_roles(request.user.id , arbitraje_id)

    if not recursos_guard(estado, rol_id):
        raise PermissionDenied

    item_active = 1
    items=validate_rol_status(estado,rol_id,item_active, arbitraje_id)

    route_conf= get_route_configuracion(estado,rol_id, arbitraje_id)
    route_seg= get_route_seguimiento(estado,rol_id)
    route_trabajos_sidebar = get_route_trabajos_sidebar(estado,rol_id,item_active)
    route_trabajos_navbar = get_route_trabajos_navbar(estado,rol_id)
    route_resultados = get_route_resultados(estado,rol_id, arbitraje_id)

    # print items
    if request.method == 'POST':
        form = CertificateToRefereeForm(request.POST, sistema_id = arbitraje.id)
        if form.is_valid():
            cert_context = create_certificate_context(arbitraje)
            for arbitro_id in form.cleaned_data['arbitros']:
                arbitro = Arbitro.objects.get(id = arbitro_id)
                now = timezone_now()
                now_date_string = '%s de %s de %s' % (now.day, MONTH_NAMES[now.month - 1], now.year)
                instance_context = {
                    'recipient_name': arbitro.nombres.split(' ')[0] + ' ' + arbitro.apellidos.split(' ')[0],
                    'roman_number': arbitraje.numero_romano,
                    'people_names': [arbitro.nombres.split(' ')[0] + ' ' + arbitro.apellidos.split(' ')[0]],
                    'peoples_id': arbitro.cedula_pasaporte
                }
                instance_context.update(cert_context)
                certificate_gen = CertificateGenerator()
                certificate = certificate_gen.get_referees_certificate(instance_context)

                name = instance_context["recipient_name"].split(' ')
                name = '_'.join(name)
                filename = "Certificado_%s_Convencion_Asovac_%s.pdf" % (name, instance_context["roman_number"])

                context_email = {
                    'sistema': arbitraje,
                    'usuario': arbitro.nombres.split(' ')[0] + ' ' + arbitro.apellidos.split(' ')[0],
                    'rol': 'árbitro de subárea'
                }
                msg_plain = render_to_string('../templates/email_templates/certificate.txt', context_email)
                msg_html = render_to_string('../templates/email_templates/certificate.html', context_email)

                email_msg = EmailMultiAlternatives('Certificado de árbitro', msg_plain, config('EMAIL_HOST_USER'), [arbitro.correo_electronico])
                email_msg.attach(filename, certificate.content, 'application/pdf')
                email_msg.attach_alternative(msg_html, "text/html")
                email_msg.send()

            messages.success(request, 'Se han enviado los certificados a los arbitros seleccionados con éxito.')
            return redirect('recursos:resources_referee')
    else:
        form = CertificateToRefereeForm(sistema_id = arbitraje.id)
    context = {
        'nombre_vista' : 'Recursos',
        'main_navbar_options' : TOP_NAVBAR_OPTIONS,
        'estado' : estado,
        'rol_id' : rol_id,
        'arbitraje_id' : arbitraje_id,
        'item_active' : item_active,
        'items':items,
        'route_conf':route_conf,
        'route_seg':route_seg,
        'route_trabajos_sidebar':route_trabajos_sidebar,
        'route_trabajos_navbar': route_trabajos_navbar,
        'route_resultados': route_resultados,
        'form': form,
        'admin_styles': True
    }

    return render(request, 'recursos_referee.html', context)


@login_required
@user_is_arbitraje
def resources_event(request):
    context = create_common_context(request)
    context['nombre_vista'] = 'Recursos'
    return render(request, 'recursos_event_certificates.html', context)


@login_required
@user_is_arbitraje
def create_logistics_certificates(request):
    arbitraje_id = request.session['arbitraje_id']
    arbitraje = Sistema_asovac.objects.get(pk=arbitraje_id)
    estado = arbitraje.estado_arbitraje
    rol_id=get_roles(request.user.id , arbitraje_id)

    if not recursos_guard(estado, rol_id):
        raise PermissionDenied

    form = MultipleRecipientsForm(request.POST or None)
    if request.method == 'POST':
        # Procesamiento manual del formulario
        recipients_tuples = list()
        n = 0
        while request.POST.get('recipients_name_' + str(n), None) or request.POST.get('recipients_email_' + str(n), None):
            name = request.POST.get('recipients_name_' + str(n), None).strip()
            email = request.POST.get('recipients_email_' + str(n), None).strip()
            recipients_tuples.append((name, email))
            n += 1
        # Validacion de campos vacios
        if len(recipients_tuples) <= 0:
            form.add_error(
                None,
                ValidationError(
                    "Por favor ingrese al menos un nombre y un correo electrónico."
                )
            )

        for r_name, r_email in recipients_tuples:
            print 'VALIDANDO TUPLAS'
            print r_name
            print r_email
            if not (r_name and r_email):
                form.add_error(
                    None,
                    ValidationError(
                        "Todos los destinatarios deben poseer un nombre y correo electrónico asociados."
                    )
                )
                break

        print 'form error', form.errors
        if not form.errors:
            arbitraje_id = request.session['arbitraje_id']
            arbitraje = Sistema_asovac.objects.get(pk=arbitraje_id)
            cert_context = create_certificate_context(arbitraje)
            certificate_gen = CertificateGenerator()
            for r_name, r_email in recipients_tuples:
                instance_context = {
                    "recipient_name": r_name,
                    "people_names": [r_name],
                }
                instance_context.update(cert_context)
                certificate = certificate_gen.get_logistics_certificate(instance_context)
                name = r_name.split(' ')
                name = '_'.join(name)
                filename = "CertificadoLogistica_%s_Convencion_Asovac_%s.pdf" % (name,
                                                                                 cert_context["roman_number"])
                context_email = {
                    'sistema': arbitraje,
                    'usuario': r_name,
                    'rol': 'miembro de la comisión de logística'
                }
                msg_plain = render_to_string('../templates/email_templates/generic_certificate.txt',
                                             context_email)
                msg_html = render_to_string('../templates/email_templates/generic_certificate.html',
                                            context_email)

                email_msg = EmailMultiAlternatives(
                    'Certificado de Comisión Logística',
                    msg_plain,
                    config('EMAIL_HOST_USER'),
                    [r_email]
                )
                email_msg.attach(filename, certificate.content, 'application/pdf')
                email_msg.attach_alternative(msg_html, "text/html")
                email_msg.send()

            messages.success(request, 'Se han enviado los certificados \
                a los destinatarios listados con éxito.')
            return redirect('recursos:create_logistics_certificates')

    context = create_common_context(request)
    context['nombre_vista'] = 'Recursos'
    context['form'] = form
    return render(request, 'recursos_event_logistics_certificate.html', context)


@login_required
@user_is_arbitraje
def resources_sesion(request):
    context = create_common_context(request)
    context['nombre_vista'] = 'Recursos'
    return render(request, 'recursos_session_certificates.html', context)


@login_required
@user_is_arbitraje
def create_session_coord_certificates(request):
    arbitraje_id = request.session['arbitraje_id']
    arbitraje = Sistema_asovac.objects.get(pk=arbitraje_id)
    estado = arbitraje.estado_arbitraje
    rol_id=get_roles(request.user.id , arbitraje_id)

    if not recursos_guard(estado, rol_id):
        raise PermissionDenied

    form = MultipleRecipientsForm(request.POST or None)
    if request.method == 'POST':
        # Procesamiento manual del formulario
        recipients_tuples = list()
        n = 0
        while request.POST.get('recipients_name_' + str(n), None) or request.POST.get('recipients_email_' + str(n), None):
            name = request.POST.get('recipients_name_' + str(n), None).strip()
            email = request.POST.get('recipients_email_' + str(n), None).strip()
            recipients_tuples.append((name, email))
            n += 1
        # Validacion de campos vacios
        if len(recipients_tuples) <= 0:
            form.add_error(
                None,
                ValidationError(
                    "Por favor ingrese al menos un nombre y un correo electrónico."
                )
            )
        for r_name, r_email in recipients_tuples:
            if not (r_name and r_email):
                form.add_error(
                    None,
                    ValidationError(
                        "Todos los destinatarios deben poseer un nombre y correo electrónico asociados."
                    )
                )
                break

        print 'form error', form.errors
        if not form.errors:
            arbitraje_id = request.session['arbitraje_id']
            arbitraje = Sistema_asovac.objects.get(pk=arbitraje_id)
            cert_context = create_certificate_context(arbitraje)
            certificate_gen = CertificateGenerator()
            for r_name, r_email in recipients_tuples:
                instance_context = {
                    "recipient_name": r_name,
                    "people_names": [r_name],
                }
                instance_context.update(cert_context)
                certificate = certificate_gen.get_session_coord_certificate(instance_context)
                name = r_name.split(' ')
                name = '_'.join(name)
                filename = "CertificadoCoordinador_de_Sesion_Convencion_Asovac_%s.pdf" % cert_context["roman_number"]
                context_email = {
                    'sistema': arbitraje,
                    'usuario': r_name,
                    'rol': 'coordinador de sesión de presentación de trabajos libres'
                }
                msg_plain = render_to_string('../templates/email_templates/generic_certificate.txt',
                                             context_email)
                msg_html = render_to_string('../templates/email_templates/generic_certificate.html',
                                            context_email)

                email_msg = EmailMultiAlternatives(
                    'Certificado de Coordinador de Sesiones',
                    msg_plain,
                    config('EMAIL_HOST_USER'),
                    [r_email]
                )
                email_msg.attach(filename, certificate.content, 'application/pdf')
                email_msg.attach_alternative(msg_html, "text/html")
                email_msg.send()

            messages.success(request, 'Se han enviado los certificados \
                a los destinatarios listados con éxito.')
            return redirect('recursos:create_session_coord_certificates')

    context = create_common_context(request)
    context['nombre_vista'] = 'Recursos'
    context['form'] = form
    return render(request, 'recursos_session_coord_certificate.html', context)


@login_required
@user_is_arbitraje
def resources_asovac(request):
    context = create_common_context(request)
    context['nombre_vista'] = 'Recursos'
    return render(request, 'recursos_asovac_documents.html', context)


@login_required
@user_is_arbitraje
def create_organizer_comitee_certificates(request):
    arbitraje_id = request.session['arbitraje_id']
    arbitraje = Sistema_asovac.objects.get(pk=arbitraje_id)
    estado = arbitraje.estado_arbitraje
    rol_id=get_roles(request.user.id , arbitraje_id)

    if not recursos_guard(estado, rol_id):
        raise PermissionDenied

    form = MultipleRecipientsForm(request.POST or None)
    if request.method == 'POST':
        # Procesamiento manual del formulario
        recipients_tuples = list()
        n = 0
        while request.POST.get('recipients_name_' + str(n), None) or request.POST.get('recipients_email_' + str(n), None):
            name = request.POST.get('recipients_name_' + str(n), None).strip()
            email = request.POST.get('recipients_email_' + str(n), None).strip()
            recipients_tuples.append((name, email))
            n += 1
        # Validacion de campos vacios
        if len(recipients_tuples) <= 0:
            form.add_error(
                None,
                ValidationError(
                    "Por favor ingrese al menos un nombre y un correo electrónico."
                )
            )
        for r_name, r_email in recipients_tuples:
            if not (r_name and r_email):
                form.add_error(
                    None,
                    ValidationError(
                        "Todos los destinatarios deben poseer un nombre y correo electrónico asociados."
                    )
                )
                break

        print 'form error', form.errors
        if not form.errors:
            arbitraje_id = request.session['arbitraje_id']
            arbitraje = Sistema_asovac.objects.get(pk=arbitraje_id)
            cert_context = create_certificate_context(arbitraje)
            certificate_gen = CertificateGenerator()
            for r_name, r_email in recipients_tuples:
                instance_context = {
                    "recipient_name": r_name,
                    "people_names": [r_name],
                }
                instance_context.update(cert_context)
                certificate = certificate_gen.get_comitee_certificate(instance_context)
                name = r_name.split(' ')
                name = '_'.join(name)
                filename = "CertificadoComite_Organizador_Convencion_Asovac_%s.pdf" % cert_context["roman_number"]
                context_email = {
                    'sistema': arbitraje,
                    'usuario': r_name,
                    'rol': 'miembro de la comisión organizadora'
                }
                msg_plain = render_to_string('../templates/email_templates/generic_certificate.txt',
                                             context_email)
                msg_html = render_to_string('../templates/email_templates/generic_certificate.html',
                                            context_email)

                email_msg = EmailMultiAlternatives(
                    'Certificado de Comisión Organizadora',
                    msg_plain,
                    config('EMAIL_HOST_USER'),
                    [r_email]
                )
                email_msg.attach(filename, certificate.content, 'application/pdf')
                email_msg.attach_alternative(msg_html, "text/html")
                email_msg.send()

            messages.success(request, 'Se han enviado los certificados \
                a los destinatarios listados con éxito.')
            return redirect('recursos:create_organizer_comitee_certificates')

    context = create_common_context(request)
    context['nombre_vista'] = 'Recursos'
    context['form'] = form
    return render(request, 'recursos_organizer_comitee_certificate.html', context)



@login_required
@user_is_arbitraje
def resources_paper(request):
    main_navbar_options = [{'title':'Configuración',   'icon': 'fa-cogs',      'active': True},
                    {'title':'Monitoreo',       'icon': 'fa-eye',       'active': False},
                    {'title':'Resultados',      'icon': 'fa-chart-area','active': False},
                    {'title':'Administración',  'icon': 'fa-archive',   'active': False}]
    # print (rol_id)
    arbitraje_id = request.session['arbitraje_id']
    arbitraje = Sistema_asovac.objects.get(pk=arbitraje_id)
    estado = arbitraje.estado_arbitraje
    rol_id=get_roles(request.user.id , arbitraje_id)

    if not recursos_guard(estado, rol_id):
        raise PermissionDenied

    item_active = 1
    items=validate_rol_status(estado,rol_id,item_active, arbitraje_id)

    route_conf= get_route_configuracion(estado,rol_id, arbitraje_id)
    route_seg= get_route_seguimiento(estado,rol_id)
    route_trabajos_sidebar = get_route_trabajos_sidebar(estado,rol_id,item_active)
    route_trabajos_navbar = get_route_trabajos_navbar(estado,rol_id)
    route_resultados = get_route_resultados(estado,rol_id, arbitraje_id)
    # print items
    if request.method == "POST":
        form = CertificateToAuthorsForm(request.POST, sistema_id = arbitraje_id)
        if form.is_valid():
            cert_context = create_certificate_context(arbitraje)
            for trabajo_id in form.cleaned_data['trabajos']:
                autores_trabajo = Autores_trabajos.objects.filter(trabajo = trabajo_id)
                """
                - context["subject_title"] = String que representa el nombre del trabajo certificado.
                - context["people_names"] = Lista de Strings con los nombres de los autores del trabajo.
                - context["roman_number"] = Numero romano de la convencion
                """
                now = timezone_now()
                now_date_string = '%s de %s de %s' % (now.day, MONTH_NAMES[now.month - 1], now.year)
                authors_emails = []
                authors_names = []
                for autor_trabajo in autores_trabajo:
                    authors_emails.append(autor_trabajo.autor.correo_electronico)
                    authors_names.append( autor_trabajo.autor.nombres.split(' ')[0] + ' ' + autor_trabajo.autor.apellidos.split(' ')[0] )
                
                instance_context = {
                    'roman_number': arbitraje.numero_romano,
                    'people_names': authors_names,
                    'subject_title': autores_trabajo.first().trabajo.titulo_espanol,
                }
                instance_context.update(cert_context)
                certificate_gen = CertificateGenerator()
                certificate = certificate_gen.get_paper_certificate(instance_context)

                filename = "Certificado_%s_Convencion_Asovac_%s.pdf" % (instance_context["subject_title"],instance_context["roman_number"])

                context_email = {
                    'sistema': arbitraje,
                    'titulo_trabajo': autores_trabajo.first().trabajo.titulo_espanol,
                    'rol': 'autor'
                }
                msg_plain = render_to_string('../templates/email_templates/certificate.txt', context_email)
                msg_html = render_to_string('../templates/email_templates/certificate.html', context_email)

                email_msg = EmailMultiAlternatives('Certificado por trabajo', msg_plain, config('EMAIL_HOST_USER'), authors_emails)
                email_msg.attach(filename, certificate.content, 'application/pdf')
                email_msg.attach_alternative(msg_html, "text/html")
                email_msg.send()

            messages.success(request, 'Se han enviado los certificados de los trabajos seleccionados al(los) autor(es) con éxito.')
            return redirect('recursos:resources_paper')
    else:    
        form = CertificateToAuthorsForm(sistema_id = arbitraje_id)
    context = {
        'nombre_vista' : 'Recursos',
        'main_navbar_options' : main_navbar_options,
        'estado' : estado,
        'rol_id' : rol_id,
        'arbitraje_id' : arbitraje_id,
        'item_active' : item_active,
        'items':items,
        'route_conf':route_conf,
        'route_seg':route_seg,
        'route_trabajos_sidebar':route_trabajos_sidebar,
        'route_trabajos_navbar': route_trabajos_navbar,
        'route_resultados': route_resultados,
        'form': form,
        'paper': True, 
        'admin_styles': True
    }
    return render(request, 'recursos_author.html', context)


@login_required
@user_is_arbitraje
def create_name_tags(request):
    arbitraje_id = request.session['arbitraje_id']
    arbitraje = Sistema_asovac.objects.get(pk=arbitraje_id)
    estado = arbitraje.estado_arbitraje
    rol_id=get_roles(request.user.id , arbitraje_id)

    if not recursos_guard(estado, rol_id):
        raise PermissionDenied

    form = MultipleRecipientsWithRoleForm(request.POST or None)
    context = create_common_context(request)
    if request.method == 'POST':
        if form.is_valid():
            # Procesamiento manual del formulario
            recipients_tuples = list()
            n = 0
            while request.POST.get('recipients_name_' + str(n), None) or request.POST.get('recipients_email_' + str(n), None):
                name = request.POST.get('recipients_name_' + str(n), None).strip()
                email = request.POST.get('recipients_email_' + str(n), None).strip()
                recipients_tuples.append((name, email))
                n += 1

            # Validacion de campos vacios
            if len(recipients_tuples) <= 0:
                form.add_error(
                    None,
                    ValidationError(
                        "Por favor ingrese al menos un nombre y un correo electrónico."
                    )
                )
            for r_name, r_email in recipients_tuples:
                if not (r_name and r_email):
                    form.add_error(
                        None,
                        ValidationError(
                            "Todos los destinatarios deben poseer un nombre y correo electrónico asociados."
                        )
                    )
                    break

            if not form.errors and form.is_valid():
                role = form.cleaned_data['role']
                arbitraje_id = request.session['arbitraje_id']
                arbitraje = Sistema_asovac.objects.get(pk=arbitraje_id)
                nametag_context = create_certificate_context(arbitraje)
                misc_gen = MiscellaneousGenerator()
                for r_name, r_email in recipients_tuples:
                    instance_context = {
                        'subject_title': r_name,
                        'role': role.capitalize(),
                    }
                    filename = "Portanombre AsoVAC.pdf"
                    instance_context.update(nametag_context)
                    certificate = misc_gen.get_name_tag(filename, instance_context)

                    context_email = {
                        'sistema': arbitraje,
                        'usuario': r_name,
                        'rol': role.capitalize()
                    }
                    msg_plain = render_to_string('../templates/email_templates/nametag.txt',
                                                 context_email)
                    msg_html = render_to_string('../templates/email_templates/nametag.html',
                                                context_email)

                    email_msg = EmailMultiAlternatives(
                        'Portanombre AsoVAC',
                        msg_plain,
                        config('EMAIL_HOST_USER'),
                        [r_email]
                    )
                    email_msg.attach(filename, certificate.content, 'application/pdf')
                    email_msg.attach_alternative(msg_html, "text/html")
                    email_msg.send()
                    print '1 EMAIL ENVIADO A:'
                    print email

                messages.success(request, 'Se han generado y enviado los portanombres con éxito')
                return redirect('recursos:create_name_tags')

    context['nombre_vista'] = 'Generación de Portanombres'
    context['form'] = form
    return render(request, 'recursos_create_nametags.html', context)


@login_required
@user_is_arbitraje
def resources_organizer(request):
    arbitraje_id = request.session['arbitraje_id']
    arbitraje = Sistema_asovac.objects.get(pk=arbitraje_id)
    estado = arbitraje.estado_arbitraje
    rol_id=get_roles(request.user.id , arbitraje_id)

    if not recursos_guard(estado, rol_id):
        raise PermissionDenied

    form = MultipleRecipientsWithDateForm(request.POST or None)
    if request.method == 'POST':
        # Procesamiento manual del formulario
        recipients_tuples = list()
        n = 0
        while request.POST.get('recipients_name_' + str(n), None) or request.POST.get('recipients_email_' + str(n), None):
            name = request.POST.get('recipients_name_' + str(n), None).strip()
            email = request.POST.get('recipients_email_' + str(n), None).strip()
            recipients_tuples.append((name, email))
            n += 1
        # Validacion de campos vacios
        if len(recipients_tuples) <= 0:
            form.add_error(
                None,
                ValidationError(
                    "Por favor ingrese al menos un nombre y un correo electrónico."
                )
            )
        for r_name, r_email in recipients_tuples:
            if not (r_name and r_email):
                form.add_error(
                    None,
                    ValidationError(
                        "Todos los destinatarios deben poseer un nombre y correo electrónico asociados."
                    )
                )
                break

        print 'form error', form.errors
        if not form.errors and form.is_valid():
            fecha_evento = form.cleaned_data['recipients_date']
            nombre_evento = form.cleaned_data['recipients_event_name']
            arbitraje_id = request.session['arbitraje_id']
            arbitraje = Sistema_asovac.objects.get(pk=arbitraje_id)
            cert_context = create_certificate_context(arbitraje)
            certificate_gen = CertificateGenerator()
            for r_name, r_email in recipients_tuples:
                instance_context = {
                    "subject_title": nombre_evento,
                    "people_names": [r_name],
                    "event_name_string": nombre_evento,
                    "event_date_string": fecha_evento.strftime('%d/%m/%Y')

                }
                instance_context.update(cert_context)
                certificate = certificate_gen.get_organizer_certificate(instance_context)
                name = r_name.split(' ')
                name = '_'.join(name)
                filename = "Certificado_Organizadores_%s_Convencion_Asovac_%s.pdf" % (name,
                                                                                 cert_context["roman_number"])
                context_email = {
                    'sistema': arbitraje,
                    'usuario': r_name,
                    'rol': 'organizador'
                }
                msg_plain = render_to_string('../templates/email_templates/generic_certificate.txt',
                                             context_email)
                msg_html = render_to_string('../templates/email_templates/generic_certificate.html',
                                            context_email)

                email_msg = EmailMultiAlternatives(
                    'Certificado de organizador',
                    msg_plain,
                    config('EMAIL_HOST_USER'),
                    [r_email]
                )
                email_msg.attach(filename, certificate.content, 'application/pdf')
                email_msg.attach_alternative(msg_html, "text/html")
                email_msg.send()

            messages.success(request, 'Se han enviado los certificados \
                a los destinatarios listados con éxito.')
            return redirect('recursos:resources_organizer')

    context = create_common_context(request)
    context['nombre_vista'] = 'Recursos'
    context['form'] = form
    context['conferencista'] = 0
    context['url_modal'] = reverse('recursos:load_organizers_modal')
    return render(request, 'recursos_organizer_certificate.html', context)


@login_required
@user_is_arbitraje
def resources_lecturer(request):
    arbitraje_id = request.session['arbitraje_id']
    arbitraje = Sistema_asovac.objects.get(pk=arbitraje_id)
    estado = arbitraje.estado_arbitraje
    rol_id=get_roles(request.user.id , arbitraje_id)

    if not recursos_guard(estado, rol_id):
        raise PermissionDenied

    form = MultipleRecipientsWithDateAndSubjectForm(request.POST or None)
    if request.method == 'POST':
        # Procesamiento manual del formulario
        recipients_tuples = list()
        n = 0
        while request.POST.get('recipients_name_' + str(n), None) or request.POST.get('recipients_email_' + str(n), None):
            name = request.POST.get('recipients_name_' + str(n), None).strip()
            email = request.POST.get('recipients_email_' + str(n), None).strip()
            recipients_tuples.append((name, email))
            n += 1
        # Validacion de campos vacios
        if len(recipients_tuples) <= 0:
            form.add_error(
                None,
                ValidationError(
                    "Por favor ingrese al menos un nombre y un correo electrónico."
                )
            )

        for r_name, r_email in recipients_tuples:
            if not (r_name and r_email):
                form.add_error(
                    None,
                    ValidationError(
                        "Todos los destinatarios deben poseer un nombre y correo electrónico asociados."
                    )
                )
                break
        if n == 0:
            form.add_error(
                None,
                ValidationError(
                    "Se requiere al menos un destinatario, por favor indique el nombre y el correo electrónico."
                )
            )
        # print 'form error', form.errors
        if not form.errors and form.is_valid():
            fecha_evento = form.cleaned_data['recipients_date']
            nombre_evento = form.cleaned_data['recipients_event_name']
            titulo_conferencia = form.cleaned_data['recipients_conference_name']
            arbitraje_id = request.session['arbitraje_id']
            arbitraje = Sistema_asovac.objects.get(pk=arbitraje_id)
            cert_context = create_certificate_context(arbitraje)
            certificate_gen = CertificateGenerator()
            for r_name, r_email in recipients_tuples:
                instance_context = {
                    "subject_title": titulo_conferencia,
                    "people_names": [r_name],
                    "event_name_string": nombre_evento,
                    "event_date_string": fecha_evento.strftime('%d/%m/%Y')

                }
                instance_context.update(cert_context)
                certificate = certificate_gen.get_lecturer_certificate(instance_context)
                name = r_name.split(' ')
                name = '_'.join(name)
                filename = "Certificado_Conferencia_%s_Convencion_Asovac_%s.pdf" % (name,
                                                                                 cert_context["roman_number"])
                context_email = {
                    'sistema': arbitraje,
                    'usuario': r_name,
                    'rol': 'conferencista'
                }
                msg_plain = render_to_string('../templates/email_templates/generic_certificate.txt',
                                             context_email)
                msg_html = render_to_string('../templates/email_templates/generic_certificate.html',
                                            context_email)

                email_msg = EmailMultiAlternatives(
                    'Certificado de conferencista',
                    msg_plain,
                    config('EMAIL_HOST_USER'),
                    [r_email]
                )
                email_msg.attach(filename, certificate.content, 'application/pdf')
                email_msg.attach_alternative(msg_html, "text/html")
                email_msg.send()

            messages.success(request, 'Se han enviado los certificados \
                a los destinatarios listados con éxito.')
            return redirect('recursos:resources_lecturer')

    context = create_common_context(request)
    context['nombre_vista'] = 'Recursos'
    context['form'] = form
    context['conferencista'] = 1
    context['url_modal'] = reverse('recursos:load_lecturers_modal')
    return render(request, 'recursos_organizer_certificate.html', context)

#---------------------------------------------------------------------------------#
#                Guarda el archivo en la carpeta del proyecto                     #
#---------------------------------------------------------------------------------#
def handle_uploaded_file(file, filename):
    if not os.path.exists('upload/'):
        os.mkdir('upload/')

    # print "Nombre del archivo a guardar: ",filename
    with open('upload/' + filename, 'wb+') as destination:
        for chunk in file.chunks():
            destination.write(chunk)

#---------------------------------------------------------------------------------#
#                Guarda el archivo en la carpeta de recursos                      #
#---------------------------------------------------------------------------------#
def save_resorce_file(file, filename):
    if not os.path.exists('static_env/media_root/memorias/'):
        os.mkdir('static_env/media_root/memorias/')

    # print "Nombre del archivo a guardar: ",filename
    with open('static_env/media_root/memorias/' + filename, 'wb+') as destination:
        for chunk in file.chunks():
            destination.write(chunk)

@login_required
@user_is_arbitraje
# Para guardar el archivo recibido del formulario
def save_file(request,type_load):

    name= str(request.FILES.get('file'))

    extension = name.split('.')
    file_name= "carga"+ type_load +"."+extension[1]
    # Permite guardar el archivo y asignarle un nombre
    handle_uploaded_file(request.FILES['file'], file_name)
    route= "upload/"+file_name
    # print "Ruta del archivo: ",route
    return route

@login_required
@user_is_arbitraje
def load_lecturers_modal(request):
    data = dict()
    arbitraje_id = request.session['arbitraje_id']
    arbitraje = Sistema_asovac.objects.get(pk=arbitraje_id)
    estado = arbitraje.estado_arbitraje
    rol_id=get_roles(request.user.id , arbitraje_id)

    if not recursos_guard(estado, rol_id):
        raise PermissionDenied

    if request.method == "POST":
        form = ImportFromExcelForm(request.POST, request.FILES)
        if form.is_valid():
			# Para guardar el archivo de forma local
            file_name= save_file(request,"conferencistas")
            extension= get_extension_file(file_name)

            #Valida el contenido del archivo
            response = validate_load_event_participants(file_name, extension,arbitraje_id, True)
            print response
            if response['status'] == 200:
                context = {
                    'title': 'Cargar Conferencistas',
                    'response': response['message'],
                    'status': response['status'],
                }
                print (response['message'], response['status'])
                data['html_form'] = render_to_string('ajax/modal_succes.html', context, request=request)
                return JsonResponse(data)
            
            else:
				messages.error(request, response['message'])


			#data['url'] = reverse('autores:authors_list')
			#data['form_is_valid'] = True
        else:       
			#data['form_is_valid'] = False
			messages.error(request, "El archivo indicado no es xls o xlsx, por favor suba un archivo en alguno de esos dos formatos.")

    else:
		form = ImportFromExcelForm()

    context = {
		'form': form,
		'rol': 'conferencistas',
		'action': reverse('recursos:load_lecturers_modal'),
        'format_url': reverse('recursos:format_import_participants', kwargs={ 'option': 2})
	}
    data['html_form'] = render_to_string('ajax/load_excel.html', context, request=request)
    return JsonResponse(data)


@login_required
@user_is_arbitraje
def load_organizers_modal(request):
    data = dict()
    arbitraje_id = request.session['arbitraje_id']
    arbitraje = Sistema_asovac.objects.get(pk=arbitraje_id)
    estado = arbitraje.estado_arbitraje
    rol_id=get_roles(request.user.id , arbitraje_id)

    if not recursos_guard(estado, rol_id):
        raise PermissionDenied

    if request.method == "POST":
        form = ImportFromExcelForm(request.POST, request.FILES)
        if form.is_valid():
			# Para guardar el archivo de forma local
            file_name= save_file(request,"organizadores")
            extension= get_extension_file(file_name)

            #Valida el contenido del archivo
            response = validate_load_event_participants(file_name, extension,arbitraje_id)
            print response

            if response['status'] == 200:
                context = {
                    'title': 'Cargar Organizadores',
                    'response': response['message'],
                    'status': response['status'],
                }
                print (response['message'], response['status'])
                data['html_form'] = render_to_string('ajax/modal_succes.html', context, request=request)
                return JsonResponse(data)

            else:
				messages.error(request, response['message'])

			#data['url'] = reverse('autores:authors_list')
			#data['form_is_valid'] = True
        else:
			#data['form_is_valid'] = False
			messages.error(request, "El archivo indicado no es xls o xlsx, por favor suba un archivo en alguno de esos dos formatos.")

    else:
		form = ImportFromExcelForm()

    context ={
		'form': form,
		'rol': 'organizadores',
		'action': reverse('recursos:load_organizers_modal'),
        'format_url': reverse('recursos:format_import_participants', kwargs={ 'option': 1})
	}
    data['html_form'] = render_to_string('ajax/load_excel.html', context, request=request)
    return JsonResponse(data)


@login_required
@user_is_arbitraje
def load_nametag_recipients_modal(request):
    data = dict()
    arbitraje_id = request.session['arbitraje_id']
    arbitraje = Sistema_asovac.objects.get(pk=arbitraje_id)
    estado = arbitraje.estado_arbitraje
    rol_id=get_roles(request.user.id , arbitraje_id)

    if not recursos_guard(estado, rol_id):
        raise PermissionDenied
        
    form = UploadFileForm(request.POST or None, request.FILES or None)
    context = create_common_context(request)
    response_context = {'title': 'Carga Destinatarios de Portanombres'}
    if request.method == "POST":
        if form.is_valid():
            extension = str(request.FILES.get('file')).decode('UTF-8').split('.')[-1]
            filename = "destinatarios_portanombre." + extension
            handle_uploaded_file(request.FILES['file'], filename)
            filename = 'upload/' + filename

            arbitraje_id = request.session['arbitraje_id']
            arbitraje = Sistema_asovac.objects.get(pk=arbitraje_id)
            nametag_context = create_certificate_context(arbitraje)

            response = validate_recipients_excel(filename, extension)

            if response['status'] == 400:
                # Se regreso algun codigo de error?
                response_context['response'] = response['message']
                response_context['status'] = 400
                data['html_form'] = render_to_string('ajax/modal_succes.html', response_context, request=request)
                return JsonResponse(data)

            sheet = xlrd.open_workbook(filename).sheet_by_index(0)
            pdf_filename = 'Portanombre_AsoVAC.pdf'
            nametag_gen = MiscellaneousGenerator()

            for fila in range(1, sheet.nrows):
                try:
                    full_name = sheet.cell_value(rowx=fila, colx=0).strip()
                    role = sheet.cell_value(rowx=fila, colx=1).strip().lower()
                    email = sheet.cell_value(rowx=fila, colx=2).strip()
                except:
                    response_context['response'] = "Ha ocurrido un error con los parámetros recibidos, vuelva a intentar."
                    response_context['status'] = 400
                    data['html_form'] = render_to_string('ajax/modal_succes.html', response_context, request=request)
                    return JsonResponse(data)

                role = role.capitalize()
                instance_context = {
                    'subject_title': full_name,
                    'role': role,
                }
                instance_context.update(nametag_context)
                nametag = nametag_gen.get_name_tag(pdf_filename, instance_context)
                context_email = {
                    'sistema': arbitraje,
                    'usuario': full_name,
                    'rol': role,
                }
                msg_plain = render_to_string('email_templates/nametag.txt', context_email)
                msg_html = render_to_string('email_templates/nametag.html', context_email)

                email_msg = EmailMultiAlternatives(
                    'Portanombre AsoVAC',
                    msg_plain,
                    config('EMAIL_HOST_USER'),
                    [email]
                )
                email_msg.attach(pdf_filename, nametag.content, 'application/pdf')
                email_msg.attach_alternative(msg_html, "text/html")
                email_msg.send()
            response_context['response'] = response['message']
            response_context['status'] = response['status']
            data['html_form'] = render_to_string('ajax/modal_succes.html', response_context, request=request)
            return JsonResponse(data)
        else:
            response_context['response'] = "El archivo indicado no es xls o xlsx, por favor suba un archivo en alguno de esos dos formatos."
            response_context['status'] = 400
            data['html_form'] = render_to_string('ajax/modal_succes.html', response_context, request=request)
            return JsonResponse(data)


    context['form'] = form
    context = {
        'form': form,
        'rol': 'destinatarios de portanombres',
        'action': reverse('recursos:load_nametag_recipients_modal'),
        'format_url': reverse('recursos:format_nametag_recipients')
    }
    data['html_form'] = render_to_string('ajax/load_excel.html', context, request=request)
    return JsonResponse(data)


def validate_load_event_participants(filename,extension,arbitraje_id, lecturer = False):
    data= dict()
    data['status']=400
    data['message']="La estructura del archivo no es correcta"
    if lecturer:
        ncols = 5
    else:
        ncols = 4
    if extension == "xlsx" or extension == "xls":
        book = xlrd.open_workbook(filename)
        if book.nsheets > 0:
            excel_file = book.sheet_by_index(0)
            email_pattern = re.compile('^([a-zA-Z0-9_\-\.]+)@([a-zA-Z0-9_\-\.]+)\.([a-zA-Z]{2,5})$')

            if excel_file.ncols == ncols:
                data['status']=200
                data['message']=""
                # se validan los campos del documento
                for fila in range(excel_file.nrows):
                    # Para no buscar los titulos
                    if fila > 0:
                        if excel_file.cell_value(rowx=fila, colx=0) == '':
							data['status']=400
							data['message'] = "Error en la fila {0} el nombre del evento es un campo obligatorio \n".format(fila)
                        

                        if excel_file.cell_value(rowx=fila, colx=1) == '':
							data['status']=400
							data['message'] = "Error en la fila {0} la fecha del evento es un campo obligatorio \n".format(fila)
                        else:
                            
                            try:
                                year, month, day, hour, minute, second = xlrd.xldate_as_tuple(excel_file.cell_value(rowx=fila, colx=1), book.datemode)
                            except:
                                data['status']=400
                                data['message'] = "Error en la fila {0} la fecha debe tener formato dd/mm/AAAA \n".format(fila)
                        
                        if excel_file.cell_value(rowx=fila, colx=2) == '':
							data['status']=400
							data['message'] = "Error en la fila {0} el correo electrónico del destinatario es un campo obligatorio \n".format(fila)
                        elif not email_pattern.match(excel_file.cell_value(rowx=fila, colx=2).strip()):
                            data['status']=400
                            data['message'] = "Error en la fila {0} el correo electrónico del destinatario no tiene un formato correcto \n".format(fila)
                        
                        if excel_file.cell_value(rowx=fila, colx=3) == '':
							data['status']=400
							data['message'] = "Error en la fila {0} el nombre del destinatario es un campo obligatorio \n".format(fila)
                        else:
							nombres_is_valid = validate_alpha(str(excel_file.cell_value(rowx=fila, colx=3)).strip())
							if not nombres_is_valid:
								data['status']=400
								data['message'] = "Error en la fila {0} el campo de nombre del destinatario, solo debe tener letras \n".format(fila)
                        
                        if lecturer:
                            if excel_file.cell_value(rowx=fila, colx=4) == '':
                                data['status']=400
                                data['message'] = "Error en la fila {0} el título de la conferencia es un campo obligatorio \n".format(fila)
    if data['status'] == 200:
        if lecturer:
            generate_massive_lecturers_certificates(book, arbitraje_id)
        else:
            generate_massive_organizer_certificates(book, arbitraje_id)
        data['message'] = "Se han generado y envíado los certificados con éxito"
        book.release_resources()
        del book
    print(data)
    return data


def generate_massive_lecturers_certificates(book, arbitraje_id):
    arbitraje = Sistema_asovac.objects.get(id = arbitraje_id)
    cert_context = create_certificate_context(arbitraje)
    certificate_gen = CertificateGenerator()
    excel_file = book.sheet_by_index(0)
    for fila in range(excel_file.nrows):
        if fila > 0:
            
            nombre_evento = excel_file.cell_value(rowx=fila, colx=0)
            year, month, day, hour, minute, second = xlrd.xldate_as_tuple(excel_file.cell_value(rowx=fila, colx=1), book.datemode)
            email = excel_file.cell_value(rowx=fila, colx=2)
            nombre = excel_file.cell_value(rowx=fila, colx=3)
            titulo_conferencia = excel_file.cell_value(rowx=fila, colx=4)
            if day < 10:
                day = '0'+ str(day)
            else:
                day = str(day)

            if month < 10:
                month = '0'+ str(month)
            else:
                month = str(month)

            instance_context = {
                "subject_title": titulo_conferencia,
                "people_names": [nombre],
                "event_name_string": nombre_evento,
                "event_date_string": day + '/'+ month +'/'+str(year)

            }
            instance_context.update(cert_context)
            certificate = certificate_gen.get_lecturer_certificate(instance_context)
            name = nombre.split(' ')
            name = '_'.join(name)
            filename = "Certificado_Conferencia_%s_Convencion_Asovac_%s.pdf" % (name,
                                                                                cert_context["roman_number"])
            context_email = {
                'sistema': arbitraje,
                'usuario': nombre,
                'rol': 'conferencista'
            }
            msg_plain = render_to_string('../templates/email_templates/generic_certificate.txt',
                                            context_email)
            msg_html = render_to_string('../templates/email_templates/generic_certificate.html',
                                        context_email)

            email_msg = EmailMultiAlternatives(
                'Certificado de conferencista',
                msg_plain,
                config('EMAIL_HOST_USER'),
                [email]
            )
            email_msg.attach(filename, certificate.content, 'application/pdf')
            email_msg.attach_alternative(msg_html, "text/html")
            email_msg.send()


def generate_massive_organizer_certificates(book, arbitraje_id):
    resultado = True
    arbitraje = Sistema_asovac.objects.get(id = arbitraje_id)
    cert_context = create_certificate_context(arbitraje)
    certificate_gen = CertificateGenerator()
    excel_file = book.sheet_by_index(0)
    for fila in range(excel_file.nrows):
        if fila > 0:
            
            nombre_evento = excel_file.cell_value(rowx=fila, colx=0)
            year, month, day, hour, minute, second = xlrd.xldate_as_tuple(excel_file.cell_value(rowx=fila, colx=1), book.datemode)
            email = excel_file.cell_value(rowx=fila, colx=2)
            nombre = excel_file.cell_value(rowx=fila, colx=3)

            if day < 10:
                day = '0'+ str(day)
            else:
                day = str(day)

            if month < 10:
                month = '0'+ str(month)
            else:
                month = str(month)

            instance_context = {
                "subject_title": nombre_evento,
                "people_names": [nombre],
                "event_name_string": nombre_evento,
                "event_date_string": day + '/' + month + '/' + str(year)

            }
            instance_context.update(cert_context)
            certificate = certificate_gen.get_organizer_certificate(instance_context)
            name = nombre_evento.split(' ')
            name = '_'.join(name)
            filename = "Certificado_Organizadores_%s_Convencion_Asovac_%s.pdf" % (name,
                                                                                cert_context["roman_number"])
            context_email = {
                'sistema': arbitraje,
                'usuario': nombre_evento,
                'rol': 'organizador'
            }
            msg_plain = render_to_string('../templates/email_templates/generic_certificate.txt',
                                            context_email)
            msg_html = render_to_string('../templates/email_templates/generic_certificate.html',
                                        context_email)

            email_msg = EmailMultiAlternatives(
                'Certificado de organizador',
                msg_plain,
                config('EMAIL_HOST_USER'),
                [email]
            )
            email_msg.attach(filename, certificate.content, 'application/pdf')
            email_msg.attach_alternative(msg_html, "text/html")
            email_msg.send()


@login_required
@user_is_arbitraje
def format_import_participants(request, option):

    arbitraje_id = request.session['arbitraje_id']
    arbitraje = Sistema_asovac.objects.get(id=arbitraje_id)
    data = []
    # Para definir propiedades del documento de excel
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=convencion_asovac_{0}_formato_conferencistas.xls'.format(arbitraje.fecha_inicio_arbitraje.year)
    workbook = xlwt.Workbook()
    worksheet = workbook.add_sheet("Conferencistas")
    date_format = xlwt.XFStyle()
    date_format.num_format_str = 'dd/MM/yyyy'
    # Para agregar los titulos de cada columna
    row_num = 0
    columns = ['Nombre Evento(*)', 'Fecha Evento(*)']
    if option == '2':
        columns.append('Correo electrónico del conferencista(*)')
        columns.append('Nombre Completo del Conferencista(*)')
        columns.append('Título de conferencia(*)')
    else:
        columns.append('Correo electrónico del organizador(*)')
        columns.append('Nombre Completo del Organizador(*)')

    for col_num in range(len(columns)):
        worksheet.write(row_num, col_num, columns[col_num])

    row_num += 1
    columns = ['Evento 1', 'Fecha', 'juanito@email.com', 'Juanito Perez']
    if option == '2':
        columns.append('Título 1')
        print(columns)

    for col_num in range(len(columns)):
        if col_num != 1:
		    worksheet.write(row_num, col_num, columns[col_num])
        else:
            worksheet.write(row_num, col_num, datetime.datetime.strptime("01/01/2019", "%d/%M/%Y").date(), date_format)

    row_num += 1
    columns = ['Evento 2',  'Fecha', 'juanita@email.com', 'Juanita Perez']
    if option == '2':
        columns.append('Título 2')
    for col_num in range(len(columns)):
        if col_num != 1:
		    worksheet.write(row_num, col_num, columns[col_num])
        else:
            worksheet.write(row_num, col_num, datetime.datetime.strptime("29/10/2019", "%d/%M/%Y").date() , date_format)

    workbook.save(response)
    return response


@login_required
@user_is_arbitraje
def format_nametag_recipients(request):
    # Para definir propiedades del documento de excel
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=convencion_asovac_formato_portanombres.xls'
    workbook = xlwt.Workbook()
    worksheet = workbook.add_sheet("Destinatarios")

    # Para agregar los titulos de cada columna
    row_num = 0
    columns = ['Nombre Completo(*)', 'Rol(*)', 'Correo electrónico(*)'] 
    for col_num in range(len(columns)):
        worksheet.write(row_num, col_num, columns[col_num])

    row_num += 1
    columns = ['Juanito Perez', 'conferencista', 'juanito@email.com']
    for col_num in range(len(columns)):
        worksheet.write(row_num, col_num, columns[col_num])

    row_num += 1
    columns = ['Juanita Perez', 'asistente', 'juanita@email.com', ]
    for col_num in range(len(columns)):
        worksheet.write(row_num, col_num, columns[col_num])

    row_num += 1
    columns = ['Juan Perez', 'autor', 'juan@email.com', ]
    for col_num in range(len(columns)):
        worksheet.write(row_num, col_num, columns[col_num])

    workbook.save(response)
    return response


# Para guardar el archivo recibido del formulario  
def save_file_memorias(requestFileName,type_load,arbitraje_id):

    # name= str(request.FILES.get('file')).decode('UTF-8')
    name= str(requestFileName).decode('UTF-8')
    file_name=''

    if type_load == "portada":   
        extension = name.split('.')        	
        file_name= "portada-"+arbitraje_id+"."+extension[1]
        # Permite guardar el archivo y asignarle un nombre
        save_resorce_file(requestFileName, file_name)
        route= "memorias/"+file_name
    else:
        if type_load == "patrocinadores":   
            extension = name.split('.')        	
            file_name= "patrocinadores-"+arbitraje_id+"."+extension[1]
            # Permite guardar el archivo y asignarle un nombre
            save_resorce_file(requestFileName, file_name)
            route= "memorias/"+file_name
        else: 
            if type_load == "afiche":  
                extension = name.split('.')        	
                file_name= "afiche-"+arbitraje_id+"."+extension[1]
                # Permite guardar el archivo y asignarle un nombre
                save_resorce_file(requestFileName, file_name)
                route= "memorias/"+file_name
    
    # print "Ruta del archivo: ",route
    return route


#---------------------------------------------------------------------------------#
#                     Vista principal para crear las memorias                     #
#---------------------------------------------------------------------------------#
@login_required
@user_is_arbitraje
def memorias(request):
    context = create_common_context(request)
    context['nombre_vista'] = 'Recursos'
    return render(request, 'recursos_asovac_memorias.html', context)


#---------------------------------------------------------------------------------#
#                             Para guardar la portada en BD                       #
#---------------------------------------------------------------------------------#
@login_required
@user_is_arbitraje
def portada(request):
    context = create_common_context(request)

    if request.method == 'POST':
        # print 'Cargar portada el metodo es post y el archivo es vliádo'
        form= UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            arbitraje_id = request.session['arbitraje_id']
            data= dict()

            # Para guardar el archivo de forma local
            requestFileName=request.FILES.get('file')
            file_name= save_file_memorias(requestFileName,"portada",arbitraje_id)
            # Para guardar ruta para la imagen de la portada
            resumen=Resumen.objects.filter(sistema_id=arbitraje_id).exists()
            if(resumen == True):
                resumen=Resumen.objects.filter(sistema_id=arbitraje_id)
                resumen.url_portada=file_name
                resumen.save()
            else:
                resumen= Resumen()
                resumen.url_portada=file_name
                resumen.sistema_id=arbitraje_id
                resumen.save()

            messages.success(request, 'La imagen fue cargada de forma correcta.')
        else:
            messages.error(request, 'Ha ocurrido un error procesando su solicitud.')
            # print pdf
            # file_name= save_file_memorias(response["Content-Disposition"],"portada",arbitraje_id)
          
            context['nombre_vista'] = 'Recursos'
            context['form'] = form
            context['arbitraje']=arbitraje_id
            return redirect('recursos:portada')
            return render(request, 'recursos_memorias_portada.html', context)
        

    form= UploadFileForm()
    context['nombre_vista'] = 'Recursos'
    context['form'] = form
    return render(request, 'recursos_memorias_portada.html', context)

#---------------------------------------------------------------------------------#
#                      Para generar vista previa de la portada                    #
#---------------------------------------------------------------------------------#
@login_required
@user_is_arbitraje
def portadaPreviewPDF(request,arbitraje):
    # Validaciones previas
    context = create_common_context(request)
    resumen=Resumen.objects.filter(sistema_id=arbitraje).exists()

    if(resumen == True):
        portada=Resumen.objects.get(sistema_id=arbitraje)
        archivo_imagen = settings.MEDIA_ROOT+'/'+portada.url_portada

    # Para generar el PDF
    response = HttpResponse(content_type='application/pdf')
    buffer = BytesIO()
    # creation of the BaseDocTempalte. showBoundary=0 to hide the debug borders
    doc = BaseDocTemplate(buffer,showBoundary=0)
    # create the frames. Here you can adjust the margins
    # Frame(leftMargin,bottomMargin,width,height,leftPadding,rightPadding,topPadding,bottomPadding,id='normal')
    frame_first_page = Frame(0, 0, 600, 843,0,0,0,0, id='first')
    # add the PageTempaltes to the BaseDocTemplate. You can also modify those to adjust the margin if you need more control over the Frames.
    doc.addPageTemplates([PageTemplate(id='first_page',frames=frame_first_page, onPage=on_first_page)])
    styles=getSampleStyleSheet()
    # start the story...
    page_content=[]
    if(resumen == True):
        img = Image(archivo_imagen,600, 843)
        page_content.append(img)
    #start the construction of the pdf
    doc.build(page_content)

    doc = buffer.getvalue()
    buffer.close()
    response.write(doc)
    return response


    # Version funcional con canvas
    # context = create_common_context(request)
    # resumen=Resumen.objects.filter(sistema_id=arbitraje).exists()

    # if(resumen == True):
    #     portada=Resumen.objects.get(sistema_id=arbitraje)
    #     archivo_imagen = settings.MEDIA_ROOT+'/'+portada.url_portada
    
    # # Para generar el pdf de la imagen de la portada
    # #Indicamos el tipo de contenido a devolver, en este caso un pdf
    # response = HttpResponse(content_type='application/pdf')
    # # para descarg el pdf
    # # response['Content-Disposition'] = 'attachment; filename="some_file.pdf"'
    # #La clase io.BytesIO permite tratar un array de bytes como un fichero binario, se utiliza como almacenamiento temporal
    # buffer = BytesIO()
    # #Canvas nos permite hacer el reporte con coordenadas X y Y
    # pdf = canvas.Canvas(buffer)
    # #Llamo al método cabecera donde están definidos los datos que aparecen en la cabecera del reporte.
    # #Utilizamos el archivo logo_django.png que está guardado en la carpeta media/imagenes
    # #Definimos el tamaño de la imagen a cargar y las coordenadas correspondientes
    # if(resumen == True):
    #     pdf.drawImage(archivo_imagen, 0,0, 595, 843,preserveAspectRatio=False)
    # #Con show page hacemos un corte de página para pasar a la siguiente
    # pdf.showPage()
    # pdf.save()
    # pdf = buffer.getvalue()
    # buffer.close()
    # response.write(pdf)
    # return response
    

#---------------------------------------------------------------------------------#
#                  Para generar sección de los patrocinadores                     #
#---------------------------------------------------------------------------------#
@login_required
@user_is_arbitraje
def patrocinadores(request):
    context = create_common_context(request)

    if request.method == 'POST':
        # print 'Cargar portada el metodo es post y el archivo es vliádo'
        form= UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            arbitraje_id = request.session['arbitraje_id']
            data= dict()

            # Para guardar el archivo de forma local
            requestFileName=request.FILES.get('file')
            file_name= save_file_memorias(requestFileName,"patrocinadores",arbitraje_id)
            # Para guardar ruta para la imagen de la portada
            resumen=Resumen.objects.filter(sistema_id=arbitraje_id).exists()
            if(resumen == True):
                resumen=Resumen.objects.get(sistema_id=arbitraje_id)
                resumen.url_patrocinadores=file_name
                resumen.save()
            else:
                resumen= Resumen()
                resumen.url_patrocinadores=file_name
                resumen.sistema_id=arbitraje_id
                resumen.save()

            messages.success(request, 'La imagen fue cargada de forma correcta.')
        else:
            messages.error(request, 'Ha ocurrido un error procesando su solicitud.')
            # print pdf
            # file_name= save_file_memorias(response["Content-Disposition"],"portada",arbitraje_id)
          
            context['nombre_vista'] = 'Recursos'
            context['form'] = form
            context['arbitraje']=arbitraje_id
            return redirect('recursos:portada')
            return render(request, 'recursos_memorias_patrocinadores.html', context)
        

    form= UploadFileForm()
    context['nombre_vista'] = 'Recursos'
    context['form'] = form
    return render(request, 'recursos_memorias_patrocinadores.html', context)

#---------------------------------------------------------------------------------#
#                Para generar la configuración de las páginas                     #
#---------------------------------------------------------------------------------#
def add_page_number(canvas, doc):
     canvas.saveState()
     canvas.setFont('Times-Roman', 10)
     page_number_text = "%d" % (doc.page)
     canvas.drawCentredString(
         0.75 * inch,
         0.75 * inch,
         page_number_text
     )
     canvas.restoreState()


def on_first_page(canvas,doc):
    canvas.saveState()
    canvas.setFont('Times-Roman',19)
    canvas.drawString(inch, 0.75 * inch, "Page %d" % doc.page)
    canvas.restoreState()

def on_remaining_pages(canvas,doc):
    canvas.saveState()
    canvas.setFont('Times-Roman',9)
    canvas.drawString(inch, 0.75 * inch, "Page %d" % doc.page)
    canvas.restoreState()

def header_footer(canvas, doc,arbitraje):
        
        sistema= Sistema_asovac.objects.get(pk=arbitraje)
        # Save the state of our canvas so we can draw on it
        canvas.saveState()
        styles = getSampleStyleSheet()

        # Estilos para el footer
        footer = getSampleStyleSheet()
        
        styleF1 = footer['Normal']
        styleF1.fontSize = 8
        styleF1.alignment = TA_LEFT
        
        styleF2 = footer['Normal']
        styleF2.fontSize = 8
        styleF2.alignment = TA_LEFT

        now = datetime.now()

        # Contenido del Header
        
        header = Image(sistema.cabecera,450, 60)
        img_table = Table(data=[[header]],colWidths=452,style=[
            # The two (0, 0) in each attribute represent the range of table cells that the style applies to. Since there's only one cell at (0, 0), it's used for both start and end of the range
            ('ALIGN', (0, 0), (0, 0), 'CENTER'),
            ('TOPPADDING',(0,0),(-1,-1),0),
            ('BOTTOMPADDING',(0,0),(-1,-1),0),
            ('BOX', (0, 0), (0, 0), 0.1, colors.HexColor('#e6e6e6')), # The fourth argument to this style attribute is the border width
            ('VALIGN', (0, 0), (0, 0), 'MIDDLE'),
            ]
        )
        w, h = header.wrap(doc.width, doc.topMargin)
        w, h = img_table.wrap(doc.width, doc.topMargin)

        # header.drawOn(canvas, doc.leftMargin, doc.height + doc.topMargin-5)
        img_table.drawOn(canvas, doc.leftMargin, doc.height + doc.topMargin-10)


        # Footer 
        contentFooter='https://www.asovac.org/blogroll-memorias-convenciones-anuales     https://www.facebook.com/ConvencionAsovac'+str(now.year)
        contentFooter2='(0212) 753-5802 convencion.asovac'+str(now.year)+'@gmail.com'
        contentFooter3=str(doc.page)
        linea= Paragraph('',styleF2)
        line = Table(data=[[linea]],colWidths=452,style=[
            # The two (0, 0) in each attribute represent the range of table cells that the style applies to. Since there's only one cell at (0, 0), it's used for both start and end of the range
            ('ALIGN', (0, 0), (0, 0), 'CENTER'),
            ('TOPPADDING',(0,0),(-1,-1),0),
            ('BOTTOMPADDING',(0,0),(-1,-1),0),
            ('BOX', (0, 0), (0, 0), 0.1, colors.HexColor('#000000')), # The fourth argument to this style attribute is the border width
            ('VALIGN', (0, 0), (0, 0), 'MIDDLE'),
            ]
        )
        
        # Contenido del Footer
        footer = Paragraph(contentFooter, styleF1)
        footer2 = Paragraph(contentFooter2, styleF2)
        pagina = Paragraph(contentFooter3, styleF2)

        w, h = line.wrap(doc.width, doc.bottomMargin)
        w, h = footer.wrap(doc.width, doc.bottomMargin)
        w, h = footer2.wrap(doc.width, doc.bottomMargin)
        w, h = pagina.wrap(doc.width, doc.bottomMargin)

        line.drawOn(canvas, doc.leftMargin, h*5.5)
        footer.drawOn(canvas, doc.leftMargin, h*4)
        footer2.drawOn(canvas, doc.leftMargin, h*3)
        pagina.drawOn(canvas, (doc.leftMargin*7)+10, h*3)
        
        # Release the canvas
        canvas.restoreState()
#---------------------------------------------------------------------------------#
#              Para generar vista previa de los patrocinadores                    #
#---------------------------------------------------------------------------------#
@login_required
@user_is_arbitraje
def patrocinadoresPreviewPDF(request,arbitraje):
    
    sistema= Sistema_asovac.objects.get(pk=arbitraje)
    resumen=Resumen.objects.filter(sistema_id=arbitraje).exists()

    if(resumen == True):
        patrocinadores=Resumen.objects.get(sistema_id=arbitraje)
        if(patrocinadores.url_patrocinadores):
            archivo_imagen = settings.MEDIA_ROOT+'/'+patrocinadores.url_patrocinadores

    response = HttpResponse(content_type='application/pdf')
    buffer = BytesIO()
  
    # creation of the BaseDocTempalte. showBoundary=0 to hide the debug borders
    doc = BaseDocTemplate(buffer,showBoundary=0,topMargin=inch+20)
    # create the frames. Here you can adjust the margins
    # Frame(leftMargin,bottomMargin,width,height,leftPadding,rightPadding,topPadding,bottomPadding,id='normal')
    frame_remaining_pages = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id='remaining')
    # add the PageTempaltes to the BaseDocTemplate. You can also modify those to adjust the margin if you need more control over the Frames.
    doc.addPageTemplates([PageTemplate(id='remaining_pages',frames=frame_remaining_pages, onPage=partial(header_footer,arbitraje=arbitraje))])
    styles=getSampleStyleSheet()
    # start the story...
    page_content=[]
    page_content.append(NextPageTemplate('remaining_pages'))  #This will load the next PageTemplate with the adjusted Frame. 
    
    page_content.append(Spacer(1,50))
    # Estilos para el titulo de la convencion
    styleNombre= styles['Heading1']
    styleNombre.fontName="Times-Roman"
    styleNombre.fontSize=30
    styleNombre.alignment=TA_CENTER

    page_content.append(Paragraph(sistema.numero_romano,styleNombre))
    page_content.append(Paragraph(sistema.nombre,styleNombre))

    # Estilos para el lema
    styleLema= styles['Heading1']
    styleLema.fontName="Times-Roman"
    styleLema.fontSize=18
    styleLema.alignment=TA_CENTER
    # styleLema.spaceBefore = 15

    page_content.append(Paragraph(sistema.slogan,styleLema))
    page_content.append(Spacer(1,100))
    
    styleTitulo= styles['Heading1']
    styleTitulo.fontName="Times-Roman"
    styleTitulo.fontSize=50
    styleTitulo.alignment=TA_CENTER
    # styleTitulo.spaceBefore = 50
    styleTitulo.textColor = colors.HexColor('#004275')

    page_content.append(Paragraph('Memorias',styleLema))

    # Para cargar los patrocinadores
    page_content.append(Spacer(1,100))
    if(resumen == True):
        if(patrocinadores.url_patrocinadores):
            img = Image(archivo_imagen,445, 200)
            page_content.append(img)

    # Estilos para el fechas
    styleFecha= styles['Heading1']
    styleFecha.fontName="Times-Roman"
    styleFecha.fontSize=12
    styleFecha.alignment=TA_CENTER
    styleTitulo.textColor = colors.HexColor('#000000')

    # Para agregar fecha de la creacion del libro
    now = datetime.now()
    fecha=str (MONTH_NAMES[now.month - 1])+' - '+str(now.year)
    page_content.append(Spacer(1,50))
    page_content.append(Paragraph(fecha,styleFecha))

    #start the construction of the pdf
    doc.build(page_content)

    doc = buffer.getvalue()
    buffer.close()
    response.write(doc)
    return response


#---------------------------------------------------------------------------------#
#                  Para generar sección de los patrocinadores                     #
#---------------------------------------------------------------------------------#
@login_required
@user_is_arbitraje
def comisionOrganizadora(request):
    context = create_common_context(request)

    if request.method == 'POST':
        # print 'Para guardar texto para la comision organizadora'
        form= resumenContentForm(request.POST)
        if form.is_valid():
            arbitraje_id = request.session['arbitraje_id']
            data= dict()

            # print request.POST['content']
            # Para guardar ruta para la imagen de la portada
            resumen=Resumen.objects.filter(sistema_id=arbitraje_id).exists()
            if (resumen == True ):
                resumen=Resumen.objects.get(sistema_id=arbitraje_id)
                resumen.comision_organizadora=request.POST['content']
                resumen.save()
            else:
                resumen= Resumen()
                resumen.comision_organizadora=request.POST['content']
                resumen.sistema_id=arbitraje_id
                resumen.save()

            messages.success(request, 'La comisión fue cargada de forma correcta.')
        else:
            messages.error(request, 'Ha ocurrido un error procesando su solicitud.')
            # print pdf
            # file_name= save_file_memorias(response["Content-Disposition"],"portada",arbitraje_id)
          
            context['nombre_vista'] = 'Recursos'
            context['form'] = form
            context['arbitraje']=arbitraje_id
            return redirect('recursos:portada')
            return render(request, 'recursos_memorias_comision_organizadora.html', context)
        

    form= resumenContentForm()
    context['nombre_vista'] = 'Recursos'
    context['form'] = form
    return render(request, 'recursos_memorias_comision_organizadora.html', context)


#---------------------------------------------------------------------------------#
#          Para generar vista previa de la comisión organizadora                  #
#---------------------------------------------------------------------------------#
@login_required
@user_is_arbitraje
def comisionPreviewPDF(request,arbitraje):
    
    sistema= Sistema_asovac.objects.get(pk=arbitraje)
    resumen=Resumen.objects.filter(sistema_id=arbitraje).exists()

    if(resumen == True):
        queryResumen=Resumen.objects.get(sistema_id=arbitraje)
        # archivo_imagen = settings.MEDIA_ROOT+'/'+patrocinadores.url_patrocinadores

    response = HttpResponse(content_type='application/pdf')
    buffer = BytesIO()
  
    # creation of the BaseDocTempalte. showBoundary=0 to hide the debug borders
    doc = BaseDocTemplate(buffer,showBoundary=0,topMargin=inch+20)
    # create the frames. Here you can adjust the margins
    # Frame(leftMargin,bottomMargin,width,height,leftPadding,rightPadding,topPadding,bottomPadding,id='normal')
    frame_remaining_pages = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id='remaining')
    # add the PageTempaltes to the BaseDocTemplate. You can also modify those to adjust the margin if you need more control over the Frames.
    doc.addPageTemplates([PageTemplate(id='remaining_pages',frames=frame_remaining_pages, onPage=partial(header_footer,arbitraje=arbitraje))])
    styles=getSampleStyleSheet()
    # start the story...
    page_content=[]
    page_content.append(NextPageTemplate('remaining_pages'))  #This will load the next PageTemplate with the adjusted Frame. 
    
    page_content.append(Spacer(1,15))

    # Estilos para el titulo de la convencion
    styleNombre= styles['Heading1']
    styleNombre.fontName="Times-Roman"
    styleNombre.fontSize=15
    styleNombre.alignment=TA_CENTER
    styleNombre.spaceAfter = 0

    page_content.append(Paragraph("<b>COMISIÓN ORGANIZADORA</b>",styleNombre))

    page_content.append(Paragraph(("<b>"+sistema.numero_romano+" "+sistema.nombre.upper()+"</b>"),styleNombre))

    # Estilos para el linea de separacion
    styleNombre= styles['Normal']
    styleNombre.fontName="Times-Roman"
    styleNombre.spaceBefore = 0
    styleNombre.spaceAfter = 15
    # Para agregar linea despues del título
    page_content.append(Paragraph("_"*87,styles['Normal']))

    # Estilos para el titulo de la convencion
    # Estilos para el titulo de la convencion
    styleComision= styles['Normal']
    styleComision.fontName="Times-Roman"
    styleComision.alignment=TA_CENTER
    styleComision.spaceBefore = 10
    styleComision.spaceAfter = 0

    
    # result = str(queryResumen.comision_organizadora).replace('\n',' <br/> ')
    # page_content.append(Paragraph(result, styleComision)) 

    # page_content.append(Paragraph(queryResumen.comision_organizadora,styleComision))
    section=""
    if(queryResumen.comision_organizadora):
        for content in queryResumen.comision_organizadora:
            section= section + content
            if (section.find("</p>") >= 0 ): 
                # print section.replace('<br />','')
                # section=section.replace('<br />','')
                page_content.append(Paragraph(section, styleComision))
                section="" 
    # page_content.append(Paragraph(queryResumen.comision_organizadora,styles['Normal']))
    # page_content.append(Spacer(1,100))
    
    #start the construction of the pdf
    doc.build(page_content)

    doc = buffer.getvalue()
    buffer.close()
    response.write(doc)
    return response



#---------------------------------------------------------------------------------#
#                  Para generar sección de las invitaciones                       #
#---------------------------------------------------------------------------------#
@login_required
@user_is_arbitraje
def invitacion(request):
    context = create_common_context(request)

    if request.method == 'POST':
        # print 'Para guardar texto para la comision organizadora'
        form= resumenContentForm(request.POST)
        if form.is_valid():
            arbitraje_id = request.session['arbitraje_id']
            data= dict()

            # Para guardar ruta para la imagen de la portada
            resumen=Resumen.objects.filter(sistema_id=arbitraje_id).exists()
            if (resumen == True ):
                resumen=Resumen.objects.get(sistema_id=arbitraje_id)
                resumen.invitacion=request.POST['content']
                resumen.save()
            else:
                resumen= Resumen()
                resumen.invitacion=request.POST['content']
                resumen.sistema_id=arbitraje_id
                resumen.save()

            messages.success(request, 'El mensaje de invitación fue cargado de forma correcta.')
        else:
            messages.error(request, 'Ha ocurrido un error procesando su solicitud.')
            # print pdf
            # file_name= save_file_memorias(response["Content-Disposition"],"portada",arbitraje_id)
          
            context['nombre_vista'] = 'Recursos'
            context['form'] = form
            context['arbitraje']=arbitraje_id
            return render(request, 'recursos_memorias_invitacion.html', context)
        

    form= resumenContentForm()
    context['nombre_vista'] = 'Recursos'
    context['form'] = form
    return render(request, 'recursos_memorias_invitacion.html', context)


#---------------------------------------------------------------------------------#
#              Para generar vista previa de las invitaciones                      #
#---------------------------------------------------------------------------------#
@login_required
@user_is_arbitraje
def invitacionPreviewPDF(request,arbitraje):
    
    sistema= Sistema_asovac.objects.get(pk=arbitraje)
    resumen=Resumen.objects.filter(sistema_id=arbitraje).exists()

    if(resumen == True):
        queryResumen=Resumen.objects.get(sistema_id=arbitraje)
        # archivo_imagen = settings.MEDIA_ROOT+'/'+patrocinadores.url_patrocinadores

    response = HttpResponse(content_type='application/pdf')
    buffer = BytesIO()
  
    # creation of the BaseDocTempalte. showBoundary=0 to hide the debug borders
    doc = BaseDocTemplate(buffer,showBoundary=0,topMargin=inch+20)
    # create the frames. Here you can adjust the margins
    # Frame(leftMargin,bottomMargin,width,height,leftPadding,rightPadding,topPadding,bottomPadding,id='normal')
    frame_remaining_pages = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id='remaining')
    # add the PageTempaltes to the BaseDocTemplate. You can also modify those to adjust the margin if you need more control over the Frames.
    doc.addPageTemplates([PageTemplate(id='remaining_pages',frames=frame_remaining_pages, onPage=partial(header_footer,arbitraje=arbitraje))])
    styles=getSampleStyleSheet()
    # start the story...
    page_content=[]
    page_content.append(NextPageTemplate('remaining_pages'))  #This will load the next PageTemplate with the adjusted Frame. 
    
    page_content.append(Spacer(1,15))

    # Estilos para el titulo de la convencion
    styleNombre= styles['Heading1']
    styleNombre.fontName="Times-Roman"
    styleNombre.fontSize=13
    styleNombre.alignment=TA_CENTER
    styleNombre.spaceAfter = 0
    
    page_content.append(Paragraph(("<b>"+sistema.numero_romano+" "+sistema.nombre.upper()+"</b>"),styleNombre))

    # Estilos para el slogan
    styleSlogan= styles['Heading1']
    styleSlogan.fontName="Times-Roman"
    styleSlogan.fontSize=17
    styleSlogan.alignment=TA_CENTER
    styleSlogan.spaceAfter = 0

    # Para generar el título 
    page_content.append(Paragraph("<b>"+sistema.slogan+"</b>",styleSlogan))

    # Estilos para las sedes y fecha de inicio y fin 
    styleSede= styles['Heading1']
    styleSede.fontName="Times-Roman"
    styleSede.fontSize=12
    styleSede.alignment=TA_CENTER
    styleSede.spaceAfter = 0
    styleSede.spaceBefore = 0
    # styleSede.leading = 20

    page_content.append(Paragraph(sistema.sedes,styleSede))
    # Para agregar fecha
    inicio_dia= datetime.strftime(sistema.fecha_inicio_arbitraje,'%d')
    fin_dia= datetime.strftime(sistema.fecha_fin_arbitraje,'%d')
    mes= datetime.strftime(sistema.fecha_fin_arbitraje,'%m')
    year= datetime.strftime(sistema.fecha_fin_arbitraje,'%Y')
   
    date_event = '%s al %s de %s de %s' % (inicio_dia,fin_dia, MONTH_NAMES[int(mes) - 1], year)
    
    page_content.append(Paragraph(date_event,styleSede))

    # Estilos para la invitación
    styleInvitacion= styles['Heading1']
    styleInvitacion.fontName="Times-Roman"
    styleInvitacion.fontSize=15
    styleInvitacion.alignment=TA_CENTER
    styleInvitacion.spaceAfter = 0

    page_content.append(Paragraph("<b>INVITACIÓN</b>",styleInvitacion))


    # Estilos para el linea de separacion
    styleNombre= styles['Normal']
    styleNombre.fontName="Times-Roman"
    styleNombre.spaceBefore = 0
    styleNombre.spaceAfter = 15
    # Para agregar linea despues del título
    page_content.append(Paragraph("_"*87,styles['Normal']))

    # Estilos para el titulo de la convencion
    # Estilos para el titulo de la convencion
    styleComision= styles['Normal']
    styleComision.fontName="Times-Roman"
    styleComision.alignment=TA_JUSTIFY
    
    # Para dar formato al texto ingresado
    section=""
    for content in queryResumen.invitacion:
        section= section + content
        if (section.find("</p>") >= 0 ): 
            # print section.replace('<br />','')
            section=section.replace('<br />','')
            page_content.append(Paragraph(section, styleComision))
            section="" 
    
    #start the construction of the pdf
    doc.build(page_content)

    doc = buffer.getvalue()
    buffer.close()
    response.write(doc)
    return response


#---------------------------------------------------------------------------------#
#                    Para generar sección de las palabras                         #
#---------------------------------------------------------------------------------#
@login_required
@user_is_arbitraje
def palabras(request):
    context = create_common_context(request)

    if request.method == 'POST':
        # print 'Para guardar texto para la comision organizadora'
        form= resumenPalabrasForm(request.POST)
        if form.is_valid():
            arbitraje_id = request.session['arbitraje_id']
            data= dict()

            # print request.POST.get('presidenteNombre')
            # Para guardar ruta para la imagen de la portada
            resumen=Resumen.objects.filter(sistema_id=arbitraje_id).exists()
            if (resumen == True ):
                resumen=Resumen.objects.get(sistema_id=arbitraje_id)
                resumen.presidente_nombre=request.POST.get('presidenteNombre')
                resumen.presidente_cargo=request.POST.get('presidenteCargo')
                resumen.presidente_contenido=request.POST.get('presidenteContenido')
                resumen.secretario_nombre=request.POST.get('secretarioNombre')
                resumen.secretario_cargo=request.POST.get('secretarioCargo')
                resumen.secretario_contenido=request.POST.get('secretarioContenido')
                resumen.cabecera=request.POST.get('cabecera')
                resumen.save()
            else:
                resumen= Resumen()
                resumen.presidente_nombre=request.POST.get('presidenteNombre')
                resumen.presidente_cargo=request.POST.get('presidenteCargo')
                resumen.presidente_contenido=request.POST.get('presidenteContenido')
                resumen.secretario_nombre=request.POST.get('secretarioNombre')
                resumen.secretario_cargo=request.POST.get('secretarioCargo')
                resumen.secretario_contenido=request.POST.get('secretarioContenido')
                resumen.sistema_id=arbitraje_id
                resumen.cabecera=request.POST.get('cabecera')
                resumen.save()

            messages.success(request, 'Las palabras de apertura del evento fueron cargadas de forma correcta.')
        else:
            messages.error(request, 'Ha ocurrido un error procesando su solicitud.')
            # print pdf
            # file_name= save_file_memorias(response["Content-Disposition"],"portada",arbitraje_id)
          
            context['nombre_vista'] = 'Recursos'
            context['form'] = form
            context['arbitraje']=arbitraje_id
            return render(request, 'recursos_memorias_invitacion.html', context)
        

    form= resumenPalabrasForm()
    context['nombre_vista'] = 'Recursos'
    context['form'] = form
    return render(request, 'recursos_memorias_palabras.html', context)

#---------------------------------------------------------------------------------#
#              Para generar vista previa de las palabras                          #
#---------------------------------------------------------------------------------#
@login_required
@user_is_arbitraje
def palabrasPreviewPDF(request,arbitraje):
    
    sistema= Sistema_asovac.objects.get(pk=arbitraje)
    resumen=Resumen.objects.filter(sistema_id=arbitraje).exists()

    if(resumen == True):
        queryResumen=Resumen.objects.get(sistema_id=arbitraje)
        # archivo_imagen = settings.MEDIA_ROOT+'/'+patrocinadores.url_patrocinadores

    response = HttpResponse(content_type='application/pdf')
    buffer = BytesIO()
  
    # creation of the BaseDocTempalte. showBoundary=0 to hide the debug borders
    doc = BaseDocTemplate(buffer,showBoundary=0,topMargin=inch+20)
    # create the frames. Here you can adjust the margins
    # Frame(leftMargin,bottomMargin,width,height,leftPadding,rightPadding,topPadding,bottomPadding,id='normal')
    frame_remaining_pages = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id='remaining')
    # add the PageTempaltes to the BaseDocTemplate. You can also modify those to adjust the margin if you need more control over the Frames.
    doc.addPageTemplates([PageTemplate(id='remaining_pages',frames=frame_remaining_pages, onPage=partial(header_footer,arbitraje=arbitraje))])
    styles=getSampleStyleSheet()
    # start the story...
    page_content=[]
    page_content.append(NextPageTemplate('remaining_pages'))  #This will load the next PageTemplate with the adjusted Frame. 
    
    # page_content.append(Spacer(1,5))

    # Para agregar cabecera personalizada
    if(queryResumen.cabecera != ""):
        # Estilos para la cabecera
        styleCabecera= styles['BodyText']
        styleCabecera.fontName="Times-Roman"
        styleCabecera.fontSize=12
        styleCabecera.alignment=TA_CENTER
        # styleCabecera.spaceAfter = 0
        # styleCabecera.spaceBefore=0 

        # Para dar formato al texto ingresado
        sectionCabecera=""
        cont=0
        # print queryResumen.cabecera
        if(queryResumen.cabecera):
            for content in queryResumen.cabecera :
                sectionCabecera= sectionCabecera + content
                if (sectionCabecera.find("</p>") >= 0 ): 
                    cont= cont+1
                    # print sectionCabecera.replace('<br />','')
                    # sectionCabecera=sectionCabecera.replace('<br />','')
                    page_content.append(Paragraph("<b>"+sectionCabecera+"</b>", styleCabecera))
                    sectionCabecera="" 
            if cont == 0:
                page_content.append(Paragraph("<b>"+queryResumen.cabecera+"</b>", styleCabecera))
        page_content.append(Spacer(1,10))
    
    # Estilos para la invitación
    styleInvitacion= styles['Heading1']
    styleInvitacion.fontName="Times-Roman"
    styleInvitacion.fontSize=15
    styleInvitacion.alignment=TA_CENTER
    styleInvitacion.spaceAfter = 0

    page_content.append(Paragraph("<b>Palabras</b>",styleInvitacion))
    if(queryResumen.presidente_nombre):
        page_content.append(Paragraph("<b>"+queryResumen.presidente_nombre+"</b>",styleInvitacion))
    if(queryResumen.presidente_cargo):
        page_content.append(Paragraph("<b>"+queryResumen.presidente_cargo+"</b>",styleInvitacion))


    # Estilos para el linea de separacion
    styleNombre= styles['Normal']
    styleNombre.fontName="Times-Roman"
    styleNombre.spaceBefore = 0
    styleNombre.spaceAfter = 15
    # Para agregar linea despues del título
    page_content.append(Paragraph("_"*87,styles['Normal']))

    # Estilos para el titulo de la convencion
    # Estilos para el titulo de la convencion
    styleComision= styles['Normal']
    styleComision.fontName="Times-Roman"
    styleComision.alignment=TA_JUSTIFY
    
    # Para dar formato al texto ingresado
    section=""
    if(queryResumen.presidente_contenido):
        for content in queryResumen.presidente_contenido:
            section= section + content
            if (section.find("</p>") >= 0 ): 
                # print section.replace('<br />','')
                section=section.replace('<br />','')
                page_content.append(Paragraph(section, styleComision))
                section="" 
    
    if queryResumen.secretario_nombre != "":
        # Para cargar informacion del secretario
        page_content.append(NextPageTemplate('remaining_pages'))
        page_content.append(PageBreak())

        # page_content.append(Spacer(1,15))
        # Para agregar cabecera personalizada
        if(queryResumen.cabecera != ""):
            # Estilos para la cabecera
            styleCabecera= styles['BodyText']
            styleCabecera.fontName="Times-Roman"
            styleCabecera.fontSize=12
            styleCabecera.alignment=TA_CENTER
            # styleCabecera.spaceAfter = 0
            # styleCabecera.spaceBefore=0 

            # Para dar formato al texto ingresado
            sectionCabecera=""
            cont=0
            # print queryResumen.cabecera
            if(queryResumen.cabecera):
                for content in queryResumen.cabecera :
                    sectionCabecera= sectionCabecera + content
                    if (sectionCabecera.find("</p>") >= 0 ): 
                        cont= cont+1
                        # print sectionCabecera.replace('<br />','')
                        # sectionCabecera=sectionCabecera.replace('<br />','')
                        page_content.append(Paragraph("<b>"+sectionCabecera+"</b>", styleCabecera))
                        sectionCabecera="" 
                if cont == 0:
                    page_content.append(Paragraph("<b>"+queryResumen.cabecera+"</b>", styleCabecera))
            page_content.append(Spacer(1,10))

        # Estilos para la invitación
        styleInvitacion= styles['Heading1']
        styleInvitacion.fontName="Times-Roman"
        styleInvitacion.fontSize=15
        styleInvitacion.alignment=TA_CENTER
        styleInvitacion.spaceAfter = 0

        page_content.append(Paragraph("<b>Palabras</b>",styleInvitacion))
        if(queryResumen.secretario_nombre):
            page_content.append(Paragraph("<b>"+queryResumen.secretario_nombre+"</b>",styleInvitacion))
        if(queryResumen.secretario_cargo):
            page_content.append(Paragraph("<b>"+queryResumen.secretario_cargo+"</b>",styleInvitacion))


        # Estilos para el linea de separacion
        styleNombre= styles['Normal']
        styleNombre.fontName="Times-Roman"
        styleNombre.spaceBefore = 0
        styleNombre.spaceAfter = 15
        # Para agregar linea despues del título
        page_content.append(Paragraph("_"*87,styles['Normal']))

        # Estilos para el titulo de la convencion
        # Estilos para el titulo de la convencion
        styleComision= styles['Normal']
        styleComision.fontName="Times-Roman"
        styleComision.alignment=TA_JUSTIFY
        
        # Para dar formato al texto ingresado
        section=""
        if(queryResumen.presidente_contenido):
            for content in queryResumen.presidente_contenido:
                section= section + content
                if (section.find("</p>") >= 0 ): 
                    # print section.replace('<br />','')
                    section=section.replace('<br />','')
                    page_content.append(Paragraph(section, styleComision))
                section="" 


    #start the construction of the pdf
    doc.build(page_content)

    doc = buffer.getvalue()
    buffer.close()
    response.write(doc)
    return response


#---------------------------------------------------------------------------------#
#                              Para guardar el afiche en BD                       #
#---------------------------------------------------------------------------------#
@login_required
@user_is_arbitraje
def afiche(request):
    context = create_common_context(request)

    if request.method == 'POST':
        # print 'Cargar portada el metodo es post y el archivo es vliádo'
        form= UploadAficheForm(request.POST, request.FILES)
        if form.is_valid():
            arbitraje_id = request.session['arbitraje_id']
            data= dict()

            # Para guardar el archivo de forma local
            requestFileName=request.FILES.get('file')
            file_name= save_file_memorias(requestFileName,"afiche",arbitraje_id)
            # Para guardar ruta para la imagen de la portada
            resumen=Resumen.objects.filter(sistema_id=arbitraje_id).exists()
            if(resumen == True):
                resumen=Resumen.objects.get(sistema_id=arbitraje_id)
                resumen.url_afiche=file_name
                resumen.afiche_titulo=request.POST.get('afiche_titulo')
                resumen.save()
            else:
                resumen= Resumen()
                resumen.afiche_titulo=file_name
                resumen.afiche_titulo=request.POST.get('afiche_titulo')
                resumen.sistema_id=arbitraje_id
                resumen.save()

            messages.success(request, 'La imagen fue cargada de forma correcta.')
        else:
            messages.error(request, 'Ha ocurrido un error procesando su solicitud.')
            # print pdf
            # file_name= save_file_memorias(response["Content-Disposition"],"portada",arbitraje_id)
          
            context['nombre_vista'] = 'Recursos'
            context['form'] = form
            context['arbitraje']=arbitraje_id
            # return redirect('recursos:portada')
            return render(request, 'recursos_memorias_afiche.html', context)
        

    form= UploadAficheForm()
    context['nombre_vista'] = 'Recursos'
    context['form'] = form
    return render(request, 'recursos_memorias_afiche.html', context)


#---------------------------------------------------------------------------------#
#                      Para generar vista previa del afiche                       #
#---------------------------------------------------------------------------------#
@login_required
@user_is_arbitraje
def afichePreviewPDF(request,arbitraje):
    # Validaciones previas
    context = create_common_context(request)
    resumen=Resumen.objects.filter(sistema_id=arbitraje).exists()

    if(resumen == True):
        afiche=Resumen.objects.get(sistema_id=arbitraje)
        if(afiche.url_afiche):
            archivo_imagen = settings.MEDIA_ROOT+'/'+afiche.url_afiche

    # Para generar el PDF
    response = HttpResponse(content_type='application/pdf')
    buffer = BytesIO()
    # creation of the BaseDocTempalte. showBoundary=0 to hide the debug borders
    doc = BaseDocTemplate(buffer,showBoundary=0)
    # create the frames. Here you can adjust the margins
    # Frame(leftMargin,bottomMargin,width,height,leftPadding,rightPadding,topPadding,bottomPadding,id='normal')
    frame_first_page =Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id='first')
    
    # add the PageTempaltes to the BaseDocTemplate. You can also modify those to adjust the margin if you need more control over the Frames.
    doc.addPageTemplates([PageTemplate(id='first_page',frames=frame_first_page)])
    styles=getSampleStyleSheet()
    # start the story...
    page_content=[]

    # Estilos para la invitación
    styleAfiche= styles['Heading1']
    styleAfiche.fontName="Times-Roman"
    styleAfiche.fontSize=12
    styleAfiche.alignment=TA_CENTER
    styleAfiche.spaceAfter = 0
    if(afiche.afiche_titulo):
        page_content.append(Paragraph(afiche.afiche_titulo,styleAfiche))
    
    if(resumen == True):
        if(afiche.url_afiche):
            img = Image(archivo_imagen,460, 655)
            # img = Image(archivo_imagen,460, 685)
            page_content.append(img)
    #start the construction of the pdf
    doc.build(page_content)

    doc = buffer.getvalue()
    buffer.close()
    response.write(doc)
    return response

#---------------------------------------------------------------------------------#
#                              Para guardar el afiche en BD                       #
#---------------------------------------------------------------------------------#
@login_required
@user_is_arbitraje
def resumenesTrabajos(request):
    context = create_common_context(request)

    if request.method == 'POST':
        # print 'Para guardar texto para la comision organizadora'
        form= comisionAcademicaForm(request.POST)
        if form.is_valid():
            arbitraje_id = request.session['arbitraje_id']
            data= dict()

            # Para guardar ruta para la imagen de la portada
            resumen=Resumen.objects.filter(sistema_id=arbitraje_id).exists()
            if (resumen == True ):
                resumen=Resumen.objects.get(sistema_id=arbitraje_id)
                resumen.coordinadores_area=request.POST['coordinadores_area']
                resumen.titulo_comision=request.POST['titulo_comision']
                resumen.miembros_comision=request.POST['miembros_comision']
                resumen.save()
            else:
                resumen= Resumen()
                resumen.coordinadores_area=request.POST['coordinadores_area']
                resumen.titulo_comision=request.POST['titulo_comision']
                resumen.miembros_comision=request.POST['miembros_comision']
                resumen.sistema_id=arbitraje_id
                resumen.save()

            messages.success(request, 'La comisión fue cargada de forma correcta.')
        else:
            messages.error(request, 'Ha ocurrido un error procesando su solicitud.')
            # print pdf
            # file_name= save_file_memorias(response["Content-Disposition"],"portada",arbitraje_id)
          
            context['nombre_vista'] = 'Recursos'
            context['form'] = form
            context['arbitraje']=arbitraje_id
            # return redirect('recursos:portada')
            return render(request, 'recursos_memorias_comision_academica.html', context)
        

    form= comisionAcademicaForm()
    context['nombre_vista'] = 'Recursos'
    context['form'] = form
    return render(request, 'recursos_memorias_comision_academica.html', context)



#---------------------------------------------------------------------------------#
#                       Para calcular los trabajos aceptados                      #
#---------------------------------------------------------------------------------#
@login_required
def getTrabajosSesion(request):

    arbitraje_id = request.session['arbitraje_id']
    group_by=' group by trab.id,trab.estatus,trab.requiere_arbitraje,trab.titulo_espanol,trab.forma_presentacion,main_a.nombre,trab.observaciones,uni.nombre,uni.facultad,uni.escuela,main_sarea.nombre'
    query= "SELECT DISTINCT(trab.id),trab.estatus,trab.palabras_clave,trab.titulo_ingles,trab.codigo,trab.requiere_arbitraje,trab.titulo_espanol,trab.forma_presentacion,trab.resumen,main_a.nombre,trab.observaciones,uni.nombre as universidad,uni.facultad,uni.escuela,STRING_AGG (distinct(concat(aut.nombres,' ',aut.apellidos)), ', ') lista_autores,main_sarea.nombre as subarea FROM trabajos_trabajo AS trab INNER JOIN autores_autores_trabajos AS aut_trab ON aut_trab.trabajo_id = trab.id INNER JOIN autores_autor AS aut ON aut.id = aut_trab.autor_id INNER JOIN main_app_sistema_asovac AS sis_aso ON sis_aso.id = aut_trab.sistema_asovac_id INNER JOIN trabajos_trabajo_subareas AS trab_suba ON trab_suba.trabajo_id = trab.id INNER JOIN main_app_sub_area AS main_sarea on main_sarea.id = trab_suba.sub_area_id INNER JOIN main_app_area AS main_a ON main_a.id= main_sarea.area_id inner join autores_universidad as uni on uni.id= aut.universidad_id "
    where=" WHERE sis_aso.id= {} AND ((trab.requiere_arbitraje = false) or (trab.padre<>0 and trab.requiere_arbitraje = false)) AND aut_trab.pagado=true AND trab.estatus='Aceptado' AND trab.confirmacion_pago='Aceptado' ".format(arbitraje_id)
    query= query+where+group_by

    data= User.objects.raw(query)
    total=0
    return data
#---------------------------------------------------------------------------------#
#          Para generar vista previa de la comisión organizadora                  #
#---------------------------------------------------------------------------------#
@login_required
@user_is_arbitraje
def resumenesPreviewPDF(request,arbitraje):
    
    sistema= Sistema_asovac.objects.get(pk=arbitraje)
    resumen=Resumen.objects.filter(sistema_id=arbitraje).exists()

    if(resumen == True):
        queryResumen=Resumen.objects.get(sistema_id=arbitraje)
        # archivo_imagen = settings.MEDIA_ROOT+'/'+patrocinadores.url_patrocinadores

    response = HttpResponse(content_type='application/pdf')
    buffer = BytesIO()
  
    # creation of the BaseDocTempalte. showBoundary=0 to hide the debug borders
    doc = BaseDocTemplate(buffer,showBoundary=0,topMargin=inch+20)
    # create the frames. Here you can adjust the margins
    # Frame(leftMargin,bottomMargin,width,height,leftPadding,rightPadding,topPadding,bottomPadding,id='normal')
    frame_resTrabLib = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id='resTrabLib')
    frame_remaining_pages = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id='remaining')
    # add the PageTempaltes to the BaseDocTemplate. You can also modify those to adjust the margin if you need more control over the Frames.
    doc.addPageTemplates([PageTemplate(id='resTrabLib',frames=frame_resTrabLib, onPage=partial(header_footer,arbitraje=arbitraje))])
    doc.addPageTemplates([PageTemplate(id='remaining_pages',frames=frame_remaining_pages, onPage=partial(header_footer,arbitraje=arbitraje))])
    styles=getSampleStyleSheet()
    # start the story...
    page_content=[]

    # Para generar título de resumenes de trabajos libres
    page_content.append(NextPageTemplate('resTrabLib'))
    
    styleArea= styles['Heading1']
    styleArea.fontName="Times-Roman"
    styleArea.fontSize=50
    styleArea.alignment=TA_RIGHT
    # styleArea.spaceBefore = 50
    styleArea.textColor = colors.HexColor('#004275')
    page_content.append(Spacer(1,250))
    page_content.append(Paragraph('<b> Resúmenes</b> <br /> <br /><b>Trabajos Libres</b>',styleArea))

    page_content.append(PageBreak())

    # Para generar la sección de comisión académica 
    page_content.append(NextPageTemplate('remaining_pages'))  #This will load the next PageTemplate with the adjusted Frame. 
    
    page_content.append(Spacer(1,0))

    # Estilos para el titulo de la convencion
    styleNombre= styles['Title']
    styleNombre.fontName="Times-Roman"
    styleNombre.fontSize=15
    styleNombre.alignment=TA_CENTER
    styleNombre.spaceAfter = 0
    styleNombre.spaceBefore = 0
    styleNombre.textColor = colors.HexColor('#000000')

    page_content.append(Paragraph("<b>COMISIÓN ACADÉMICA</b>",styleNombre))

     # Estilos para los coordinadores de la comision
    styleCoordinadores= styles['Title']
    styleCoordinadores.fontName="Times-Roman"
    styleCoordinadores.fontSize=13
    styleCoordinadores.alignment=TA_CENTER
    styleCoordinadores.spaceAfter = 0
    page_content.append(Paragraph("<b>Coordinadores por Área</b>",styleCoordinadores))

    # Estilos para el titulo de la convencion
    # Estilos para el titulo de la convencion
    styleComision= styles['Normal']
    styleComision.fontName="Times-Roman"
    styleComision.alignment=TA_JUSTIFY
    styleComision.spaceBefore = 10
    styleComision.spaceAfter = 0

    section=""
    if(queryResumen.coordinadores_area):
        for content in queryResumen.coordinadores_area:
            section= section + content
            if (section.find("</p>") >= 0 ): 
                # print section.replace('<br />','')
                # section=section.replace('<br />','')
                page_content.append(Paragraph(section, styleComision))
                section="" 
    
     # Estilos para el titulo de los miembros de la comision
    styleMiembros= styles['Title']
    styleMiembros.fontName="Times-Roman"
    styleMiembros.fontSize=13
    styleMiembros.alignment=TA_CENTER
    styleMiembros.spaceAfter = 0
    
    if(queryResumen.titulo_comision):
        page_content.append(Paragraph("<b>"+queryResumen.titulo_comision+"</b>",styleMiembros))

    section=""
    if(queryResumen.miembros_comision):
        for content in queryResumen.miembros_comision:
            section= section + content
            if (section.find("</p>") >= 0 ): 
                # print section.replace('<br />','')
                # section=section.replace('<br />','')
                page_content.append(Paragraph(section, styleComision))
                section="" 

    # Para obtener lsitado de areas, subareas y trabajos aceptados
    areas_list = list (Area.objects.all().order_by('nombre'))
    subAreas_list = list (Sub_area.objects.all().order_by('nombre'))
    trabajos_list=getTrabajosSesion(request)
    trabajos_array=[]
    areas_check=[]
    subareas_check=[]
    trabajos_check=[]
    # Para generar arreglo con el resultado de la consulta
    
    for trabajo in trabajos_list:
        trabajos_array.append(trabajo)


    # areas_list.pop("test")
    # Para recorrer el listado de areas, subareas y cargar la información de cada trabajo
    for area in areas_list:
        # print area.nombre
        for subarea in subAreas_list:
            # print subarea.nombre
            for trabajo in trabajos_array:
    
                if (subarea.nombre == trabajo.subarea) and (area.nombre == trabajo.nombre):
                    # Para agregar página con el título del área de los trabajos a cargar
                    if (area.nombre not in areas_check):
                        # print "area sin cargar {}".format(area.nombre)
                        areas_check.append(area.nombre)
                        # Para cargar nueva pagina
                        page_content.append(NextPageTemplate('remaining_pages'))
                        page_content.append(PageBreak())
                        
                        styleArea= styles['Heading1']
                        styleArea.fontName="Times-Roman"
                        styleArea.fontSize=50
                        styleArea.alignment=TA_RIGHT
                        # styleArea.spaceBefore = 50
                        styleArea.textColor = colors.HexColor('#004275')
                        page_content.append(Spacer(1,250))
                        page_content.append(Paragraph('<b> Área</b> <br /> <br /><b>'+area.nombre+'</b>',styleArea))

                    
                    
                    # Para cargar nueva pagina
                    page_content.append(NextPageTemplate('remaining_pages'))
                    page_content.append(PageBreak())
                    # # Para agregar título de cada subarea
                    if (subarea.nombre not in subareas_check):
                        subareas_check.append(subarea.nombre)
                        # Estilos para las subáreas
                        styleSubAreas= styles['Title']
                        styleSubAreas.fontName="Times-Roman"
                        styleSubAreas.fontSize=19
                        styleSubAreas.alignment=TA_CENTER
                        styleSubAreas.spaceAfter = 0
                        styleSubAreas.spaceBefore = 0
                        styleSubAreas.textColor = colors.HexColor('#000000')
                        page_content.append(Paragraph("Sesión <b> "+subarea.nombre+"</b>",styleSubAreas))
                        # Estilos para el linea de separacion
                        styleNombre= styles['Normal']
                        styleNombre.fontName="Times-Roman"
                        # styleNombre.spaceBefore = 0
                        # styleNombre.spaceAfter = 0
                        # Para agregar linea despues del título
                        page_content.append(Paragraph("_"*87,styles['Normal']))
                        page_content.append(Spacer(1,15))

                    # Para cargar contenido de la pagina

                    styleResumen= styles['Title']
                    styleResumen.fontName="Times-Roman"
                    styleResumen.fontSize=9
                    styleResumen.alignment=TA_CENTER
                    styleResumen.spaceAfter = 0
                    styleResumen.spaceBefore = 0

                    page_content.append(Paragraph("<b>("+trabajo.codigo+") "+trabajo.titulo_espanol.upper()+"</b>",styleResumen))
                    if trabajo.titulo_ingles != "":
                        page_content.append(Paragraph("<b>("+trabajo.titulo_ingles+")</b>",styleResumen))
                    page_content.append(Paragraph(trabajo.lista_autores,styleResumen))
                    page_content.append(Paragraph(trabajo.universidad+" "+trabajo.facultad+" "+trabajo.escuela,styleResumen))
                    
                    page_content.append(Spacer(1,15))
                    styleContent= styles['Normal']
                    styleContent.fontName="Times-Roman"
                    styleContent.alignment=TA_JUSTIFY
                    styleContent.spaceBefore = 0
                    styleContent.spaceAfter = 0
                    page_content.append(Paragraph(trabajo.resumen,styleContent))
                    if trabajo.palabras_clave != "":
                        page_content.append(Spacer(1,15))
                        page_content.append(Paragraph("<b>Palabras clave: </b>"+trabajo.palabras_clave,styleContent))

                    trabajos_check.append(trabajo)
                
            
            # eliminar trabajos ya asignados para reducir el recorrido del ciclo for
            for check in trabajos_check:
                trabajos_array.remove(check) 
            trabajos_check=[] 

    #start the construction of the pdf
    doc.build(page_content)

    doc = buffer.getvalue()
    buffer.close()
    response.write(doc)
    return response


#---------------------------------------------------------------------------------#
#              Para generar vista previa de seccion del indice                    #
#---------------------------------------------------------------------------------#

@login_required
@user_is_arbitraje
def memoriasIndice(request,arbitraje):
    
    sistema= Sistema_asovac.objects.get(pk=arbitraje)
    resumen=Resumen.objects.filter(sistema_id=arbitraje).exists()

    if(resumen == True):
        queryResumen=Resumen.objects.get(sistema_id=arbitraje)
        # archivo_imagen = settings.MEDIA_ROOT+'/'+patrocinadores.url_patrocinadores

    response = HttpResponse(content_type='application/pdf')
    buffer = BytesIO()
  
    # creation of the BaseDocTempalte. showBoundary=0 to hide the debug borders
    doc = BaseDocTemplate(buffer,showBoundary=0,topMargin=inch+20)
    # create the frames. Here you can adjust the margins
    # Frame(leftMargin,bottomMargin,width,height,leftPadding,rightPadding,topPadding,bottomPadding,id='normal')
    frame_resTrabLib = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id='resTrabLib')
    frame_remaining_pages = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id='remaining')
    # add the PageTempaltes to the BaseDocTemplate. You can also modify those to adjust the margin if you need more control over the Frames.
    doc.addPageTemplates([PageTemplate(id='resTrabLib',frames=frame_resTrabLib, onPage=partial(header_footer,arbitraje=arbitraje))])
    doc.addPageTemplates([PageTemplate(id='remaining_pages',frames=frame_remaining_pages, onPage=partial(header_footer,arbitraje=arbitraje))])
    styles=getSampleStyleSheet()
    # start the story...
    page_content=[]

    
    page_content.append(Spacer(1,0))

    section=""

    # Para obtener lsitado de areas, subareas y trabajos aceptados
    areas_list = list (Area.objects.all().order_by('nombre'))
    subAreas_list = list (Sub_area.objects.all().order_by('nombre'))
    trabajos_list=getTrabajosSesion(request)
    trabajos_array=[]
    areas_check=[]
    subareas_check=[]
    trabajos_check=[]
    # Para generar arreglo con el resultado de la consulta
    
    for trabajo in trabajos_list:
        trabajos_array.append(trabajo)


    # areas_list.pop("test")
    # Para recorrer el listado de areas, subareas y cargar la información de cada trabajo
    for area in areas_list:
        # print area.nombre
        for subarea in subAreas_list:
            # print subarea.nombre
            for trabajo in trabajos_array:
    
                if (subarea.nombre == trabajo.subarea) and (area.nombre == trabajo.nombre):
                    # Para agregar página con el título del área de los trabajos a cargar
                    if (area.nombre not in areas_check):
                        # print "area sin cargar {}".format(area.nombre)
                        areas_check.append(area.nombre)
                        # Para cargar nueva pagina
                        page_content.append(NextPageTemplate('remaining_pages'))
                        page_content.append(PageBreak())
                        
                        styleArea= styles['Heading1']
                        styleArea.fontName="Times-Roman"
                        styleArea.fontSize=50
                        styleArea.alignment=TA_RIGHT
                        # styleArea.spaceBefore = 50
                        styleArea.textColor = colors.HexColor('#004275')
                        page_content.append(Spacer(1,250))
                        page_content.append(Paragraph('<b> Área</b> <br /> <br /><b>'+area.nombre+'</b>',styleArea))

                    
                    
                    # Para cargar nueva pagina
                    page_content.append(NextPageTemplate('remaining_pages'))
                    page_content.append(PageBreak())
                    # # Para agregar título de cada subarea
                    if (subarea.nombre not in subareas_check):
                        subareas_check.append(subarea.nombre)
                        # Estilos para las subáreas
                        styleSubAreas= styles['Title']
                        styleSubAreas.fontName="Times-Roman"
                        styleSubAreas.fontSize=19
                        styleSubAreas.alignment=TA_CENTER
                        styleSubAreas.spaceAfter = 0
                        styleSubAreas.spaceBefore = 0
                        styleSubAreas.textColor = colors.HexColor('#000000')
                        page_content.append(Paragraph("Sesión <b> "+subarea.nombre+"</b>",styleSubAreas))
                        # Estilos para el linea de separacion
                        styleNombre= styles['Normal']
                        styleNombre.fontName="Times-Roman"
                        # styleNombre.spaceBefore = 0
                        # styleNombre.spaceAfter = 0
                        # Para agregar linea despues del título
                        page_content.append(Paragraph("_"*87,styles['Normal']))
                        page_content.append(Spacer(1,15))

                    # Para cargar contenido de la pagina

                    styleResumen= styles['Title']
                    styleResumen.fontName="Times-Roman"
                    styleResumen.fontSize=9
                    styleResumen.alignment=TA_CENTER
                    styleResumen.spaceAfter = 0
                    styleResumen.spaceBefore = 0

                    page_content.append(Paragraph("<b>("+trabajo.codigo+") "+trabajo.titulo_espanol.upper()+"</b>",styleResumen))
                    if trabajo.titulo_ingles != "":
                        page_content.append(Paragraph("<b>("+trabajo.titulo_ingles+")</b>",styleResumen))
                    page_content.append(Paragraph(trabajo.lista_autores,styleResumen))
                    page_content.append(Paragraph(trabajo.universidad+" "+trabajo.facultad+" "+trabajo.escuela,styleResumen))
                    
                    page_content.append(Spacer(1,15))
                    styleContent= styles['Normal']
                    styleContent.fontName="Times-Roman"
                    styleContent.alignment=TA_JUSTIFY
                    styleContent.spaceBefore = 0
                    styleContent.spaceAfter = 0
                    page_content.append(Paragraph(trabajo.resumen,styleContent))
                    if trabajo.palabras_clave != "":
                        page_content.append(Spacer(1,15))
                        page_content.append(Paragraph("<b>Palabras clave: </b>"+trabajo.palabras_clave,styleContent))

                    trabajos_check.append(trabajo)
                
            
            # eliminar trabajos ya asignados para reducir el recorrido del ciclo for
            for check in trabajos_check:
                trabajos_array.remove(check) 
            trabajos_check=[] 

    #start the construction of the pdf
    doc.build(page_content)

    doc = buffer.getvalue()
    buffer.close()
    response.write(doc)
    return response
#---------------------------------------------------------------------------------#
#              Para generar vista previa de seccion de eventos                    #
#---------------------------------------------------------------------------------#
@login_required
@user_is_arbitraje
def eventosPreviewPDF(request,arbitraje):
    
    sistema= Sistema_asovac.objects.get(pk=arbitraje)
    eventos=Evento.objects.filter(sistema_asovac=arbitraje).exists()

    if(eventos == True):
        queryEventos=Evento.objects.filter(sistema_asovac=arbitraje).order_by('fecha_inicio','hora_inicio')
    else:
        queryEventos=None 
    response = HttpResponse(content_type='application/pdf')
    buffer = BytesIO()
  
    # creation of the BaseDocTempalte. showBoundary=0 to hide the debug borders
    doc = BaseDocTemplate(buffer,showBoundary=0,topMargin=inch+20)
    # create the frames. Here you can adjust the margins
    # Frame(leftMargin,bottomMargin,width,height,leftPadding,rightPadding,topPadding,bottomPadding,id='normal')
    frame_resTrabLib = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id='resTrabLib')
    frame_remaining_pages = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id='remaining')
    # add the PageTempaltes to the BaseDocTemplate. You can also modify those to adjust the margin if you need more control over the Frames.
    doc.addPageTemplates([PageTemplate(id='resTrabLib',frames=frame_resTrabLib, onPage=partial(header_footer,arbitraje=arbitraje))])
    doc.addPageTemplates([PageTemplate(id='remaining_pages',frames=frame_remaining_pages, onPage=partial(header_footer,arbitraje=arbitraje))])
    styles=getSampleStyleSheet()
    # start the story...
    page_content=[]

    # Para generar título de resumenes de trabajos libres
    page_content.append(NextPageTemplate('resTrabLib'))
    
    styleArea= styles['Heading1']
    styleArea.fontName="Times-Roman"
    styleArea.fontSize=50
    styleArea.alignment=TA_RIGHT
    # styleArea.spaceBefore = 50
    styleArea.textColor = colors.HexColor('#004275')
    page_content.append(Spacer(1,250))
    page_content.append(Paragraph('<b> Programa de </b> <br /> <br /><b> Eventos</b> <br /> <br /><b> Especiales</b>',styleArea))

    page_content.append(PageBreak())

    # Para generar la sección de comisión académica 
    page_content.append(NextPageTemplate('remaining_pages'))  #This will load the next PageTemplate with the adjusted Frame. 
    
    page_content.append(Spacer(1,0))

    # Estilos para el titulo de los eventos
    styleNombre= styles['Title']
    styleNombre.fontName="Times-Roman"
    styleNombre.fontSize=20
    styleNombre.alignment=TA_CENTER
    styleNombre.spaceAfter = 0
    styleNombre.spaceBefore = 0
    styleNombre.textColor = colors.HexColor('#000000')

    page_content.append(Paragraph("<b>Programa General de Eventos Especiales</b>",styleNombre))
    page_content.append(Spacer(1,20))
    # print queryEventos[0].fecha_inicio
    # print queryEventos[0].fecha_inicio.weekday()
    # print DAY_NAMES[queryEventos[0].fecha_inicio.weekday()]
    day_before=''

    # Estilos para el titulo para la fecha de los eventos
    styleTitleTable= styles['BodyText']
    styleTitleTable.fontName="Times-Roman"
    styleTitleTable.fontSize=10
    styleTitleTable.alignment=TA_JUSTIFY
    styleTitleTable.spaceAfter = 0
    styleTitleTable.spaceBefore = 0
    styleTitleTable.textColor = colors.HexColor('#ffffff')

    styleContentTable= styles['Normal']
    styleContentTable.fontName="Times-Roman"
    styleContentTable.fontSize=10
    styleContentTable.alignment=TA_JUSTIFY
    styleContentTable.spaceAfter = 0
    styleContentTable.spaceBefore = 0

    table_title=[]
    table_content=[]
    contItem=0
    has_title=False
    if(queryEventos != None):
        for event in queryEventos:
            actual_weekday= DAY_NAMES[event.fecha_inicio.weekday()]
            actual_day= str (event.fecha_inicio.day)
            actual_month= MONTH_NAMES[event.fecha_inicio.month-1]
            
            # print event.locacion_evento.lugar
            if (contItem+1) < len(queryEventos):
                day_next=queryEventos[contItem+1].fecha_inicio
                
            else:
                day_next=''
            
            if day_next == event.fecha_inicio:
                # Para agregar el titulo al inicio de la tabla
                if has_title == False:
                    # Para cargar el título de la tabla
                    t1 = Paragraph(('HORARIO').encode('utf-8'), styleTitleTable)
                    t2 = Paragraph('UBICACIÓN'.encode('utf-8'), styleTitleTable)
                    t3 = Paragraph('DESCRIPCIÓN'.encode('utf-8'), styleTitleTable)
                    table_content.append([t1, t2, t3])
                    has_title=True
                
                horario= event.hora_inicio.strftime(" %I:%M%p")+'-'+event.hora_fin.strftime(" %I:%M%p")
                horario= Paragraph((horario).encode('utf-8'), styleContentTable)
                ubicacion=Paragraph((event.locacion_evento.lugar).encode('utf-8'), styleContentTable)
                descripcion=Paragraph((event.descripcion).encode('utf-8'), styleContentTable)

                table_content.append([horario,ubicacion,descripcion])
            else:
                page_content.append(Paragraph(actual_weekday+" "+actual_day+" de "+actual_month,styles['Normal']))
                page_content.append(Spacer(1,5))

                # Para agregar el titulo al inicio de la tabla
                if has_title == False:
                    # Para cargar el título de la tabla
                    t1 = Paragraph(('HORARIO').encode('utf-8'), styleTitleTable)
                    t2 = Paragraph('UBICACIÓN'.encode('utf-8'), styleTitleTable)
                    t3 = Paragraph('DESCRIPCIÓN'.encode('utf-8'), styleTitleTable)
                    table_content.append([t1, t2, t3])

                horario= event.hora_inicio.strftime(" %I:%M%p")+'-'+event.hora_fin.strftime(" %I:%M%p")
                horario=Paragraph((horario).encode('utf-8'), styleContentTable)
                ubicacion=Paragraph((event.locacion_evento.lugar).encode('utf-8'), styleContentTable)
                descripcion=Paragraph((event.descripcion).encode('utf-8'), styleContentTable)

                table_content.append([horario,ubicacion,descripcion])

                tabla = Table(data = table_content,colWidths=(150,150,150),
                            style = [
                                    ('GRID',(0,0),(-1,-1),0.5,colors.HexColor('#cce0ff')),
                                    ('ALIGN', (1,1), (-1,-1), 'RIGHT'),
                                    ('BOX',(0,0),(-1,-1),2,colors.white),
                                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#99c2ff')),
                                    ]
                            )
                
                page_content.append(tabla)
                page_content.append(Spacer(1,10))
                table_content=[]
                has_title= False

            
            # day_before=DAY_NAMES[event.fecha_inicio.weekday()]
            day_before=event.fecha_inicio
            contItem=contItem+1
            # Para agregar contenido de la tabla

    #start the construction of the pdf
    doc.build(page_content)

    doc = buffer.getvalue()
    buffer.close()
    response.write(doc)
    return response


#---------------------------------------------------------------------------------#
#            Para generar vista previa para el resumen de la convención           #
#---------------------------------------------------------------------------------#
@login_required
@user_is_arbitraje
def memoriasPreviewPDF(request,arbitraje):
    # Validaciones previas
    context = create_common_context(request)
    resumen=Resumen.objects.filter(sistema_id=arbitraje).exists()

    if(resumen == True):
        portada=Resumen.objects.get(sistema_id=arbitraje)
        if(portada.url_portada):
            archivo_imagen = settings.MEDIA_ROOT+'/'+portada.url_portada
    else:
        # Para generar el PDF
        response = HttpResponse(content_type='application/pdf')
        buffer = BytesIO()
        # creation of the BaseDocTempalte. showBoundary=0 to hide the debug borders
        doc = BaseDocTemplate(buffer,showBoundary=0)
        # create the frames. Here you can adjust the margins
        # Frame(leftMargin,bottomMargin,width,height,leftPadding,rightPadding,topPadding,bottomPadding,id='normal')
        frame_first_page = Frame(0, 0, 600, 843,0,0,0,0, id='first')
        # add the PageTempaltes to the BaseDocTemplate. You can also modify those to adjust the margin if you need more control over the Frames.
        doc.addPageTemplates([PageTemplate(id='first_page',frames=frame_first_page)])
        styles=getSampleStyleSheet()
        # start the story...
        page_content=[]
        page_content.append(NextPageTemplate('first_page'))
        #start the construction of the pdf
        doc.build(page_content)
        doc = buffer.getvalue()
        buffer.close()
        response.write(doc)
        return response

    response = HttpResponse(content_type='application/pdf')
    buffer = BytesIO()
    # creation of the BaseDocTempalte. showBoundary=0 to hide the debug borders
    doc = BaseDocTemplate(buffer,showBoundary=0)
    # create the frames. Here you can adjust the margins
    # Frame(leftMargin,bottomMargin,width,height,leftPadding,rightPadding,topPadding,bottomPadding,id='normal')
    frame_first_page = Frame(0, 0, 600, 843,0,0,0,0, id='first')
    # add the PageTempaltes to the BaseDocTemplate. You can also modify those to adjust the margin if you need more control over the Frames.
    doc.addPageTemplates([PageTemplate(id='first_page',frames=frame_first_page, onPage=on_first_page)])
    styles=getSampleStyleSheet()
    # start the story...
    page_content=[]
    ############################## Para generar la portada #################################
    
    if(resumen == True):
        img = Image(archivo_imagen,600, 843)
        page_content.append(img)

    ############################## Para cargar patrocinadores ##############################
    sistema= Sistema_asovac.objects.get(pk=arbitraje)
    resumen=Resumen.objects.filter(sistema_id=arbitraje).exists()

    if(resumen == True):
        patrocinadores=Resumen.objects.get(sistema_id=arbitraje)
        if(patrocinadores.url_patrocinadores):
            archivo_imagen = settings.MEDIA_ROOT+'/'+patrocinadores.url_patrocinadores

    # Frame(leftMargin,bottomMargin,width,height,leftPadding,rightPadding,topPadding,bottomPadding,id='normal')
    frame_remaining_pages = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id='remaining')
    # add the PageTempaltes to the BaseDocTemplate. You can also modify those to adjust the margin if you need more control over the Frames.
    doc.addPageTemplates([PageTemplate(id='remaining_pages',frames=frame_remaining_pages, onPage=partial(header_footer,arbitraje=arbitraje))])
    styles=getSampleStyleSheet()
    # start the story...
    page_content.append(NextPageTemplate('remaining_pages'))  #This will load the next PageTemplate with the adjusted Frame. 
    
    page_content.append(Spacer(1,50))
    # Estilos para el titulo de la convencion
    styleNombre= styles['Heading1']
    styleNombre.fontName="Times-Roman"
    styleNombre.fontSize=30
    styleNombre.alignment=TA_CENTER

    page_content.append(Paragraph(sistema.numero_romano,styleNombre))
    page_content.append(Paragraph(sistema.nombre,styleNombre))

    # Estilos para el lema
    styleLema= styles['Heading1']
    styleLema.fontName="Times-Roman"
    styleLema.fontSize=18
    styleLema.alignment=TA_CENTER
    # styleLema.spaceBefore = 15

    page_content.append(Paragraph(sistema.slogan,styleLema))
    page_content.append(Spacer(1,100))
    
    styleTitulo= styles['Heading1']
    styleTitulo.fontName="Times-Roman"
    styleTitulo.fontSize=50
    styleTitulo.alignment=TA_CENTER
    # styleTitulo.spaceBefore = 50
    styleTitulo.textColor = colors.HexColor('#004275')

    page_content.append(Paragraph('Memorias',styleLema))

    # Para cargar los patrocinadores
    page_content.append(Spacer(1,100))
    if(resumen == True):
        img = Image(archivo_imagen,445, 200)
        page_content.append(img)

    # Estilos para el fechas
    styleFecha= styles['Heading1']
    styleFecha.fontName="Times-Roman"
    styleFecha.fontSize=12
    styleFecha.alignment=TA_CENTER
    styleTitulo.textColor = colors.HexColor('#000000')

    # Para agregar fecha de la creacion del libro
    now = datetime.now()
    fecha=str (MONTH_NAMES[now.month - 1])+' - '+str(now.year)
    page_content.append(Spacer(1,50))
    page_content.append(Paragraph(fecha,styleFecha))

    ############################## Comisión Organizadora ###################################
    page_content.append(PageBreak())
    sistema= Sistema_asovac.objects.get(pk=arbitraje)
    resumen=Resumen.objects.filter(sistema_id=arbitraje).exists()

    if(resumen == True):
        queryResumen=Resumen.objects.get(sistema_id=arbitraje)
        # archivo_imagen = settings.MEDIA_ROOT+'/'+patrocinadores.url_patrocinadores
  
    styles=getSampleStyleSheet()
    # start the story...
    page_content.append(NextPageTemplate('remaining_pages'))  #This will load the next PageTemplate with the adjusted Frame. 
    
    page_content.append(Spacer(1,46))

    # Estilos para el titulo de la convencion
    styleNombre= styles['Heading1']
    styleNombre.fontName="Times-Roman"
    styleNombre.fontSize=15
    styleNombre.alignment=TA_CENTER
    styleNombre.spaceAfter = 0

    page_content.append(Paragraph("<b>COMISIÓN ORGANIZADORA</b>",styleNombre))

    page_content.append(Paragraph(("<b>"+sistema.numero_romano+" "+sistema.nombre.upper()+"</b>"),styleNombre))

    # Estilos para el linea de separacion
    styleNombre= styles['Normal']
    styleNombre.fontName="Times-Roman"
    styleNombre.spaceBefore = 0
    styleNombre.spaceAfter = 15
    # Para agregar linea despues del título
    page_content.append(Paragraph("_"*87,styles['Normal']))

    # Estilos para el titulo de la convencion
    # Estilos para el titulo de la convencion
    styleComision= styles['Normal']
    styleComision.fontName="Times-Roman"
    styleComision.alignment=TA_CENTER
    styleComision.spaceBefore = 10
    styleComision.spaceAfter = 0

    
    # result = str(queryResumen.comision_organizadora).replace('\n',' <br/> ')
    # page_content.append(Paragraph(result, styleComision)) 

    # page_content.append(Paragraph(queryResumen.comision_organizadora,styleComision))
    section=""
    if(queryResumen.comision_organizadora):
        for content in queryResumen.comision_organizadora:
            section= section + content
            if (section.find("</p>") >= 0 ): 
                # print section.replace('<br />','')
                # section=section.replace('<br />','')
                page_content.append(Paragraph(section, styleComision))
                section="" 


    ##############################      Invitación       ###################################
    page_content.append(PageBreak())

    sistema= Sistema_asovac.objects.get(pk=arbitraje)
    resumen=Resumen.objects.filter(sistema_id=arbitraje).exists()

    if(resumen == True):
        queryResumen=Resumen.objects.get(sistema_id=arbitraje)

    styles=getSampleStyleSheet()
    page_content.append(NextPageTemplate('remaining_pages'))  #This will load the next PageTemplate with the adjusted Frame. 
    
    page_content.append(Spacer(1,15))

    # Estilos para el titulo de la convencion
    styleNombre= styles['Heading1']
    styleNombre.fontName="Times-Roman"
    styleNombre.fontSize=13
    styleNombre.alignment=TA_CENTER
    styleNombre.spaceAfter = 0
    
    page_content.append(Paragraph(("<b>"+sistema.numero_romano+" "+sistema.nombre.upper()+"</b>"),styleNombre))

    # Estilos para el slogan
    styleSlogan= styles['Heading1']
    styleSlogan.fontName="Times-Roman"
    styleSlogan.fontSize=17
    styleSlogan.alignment=TA_CENTER
    styleSlogan.spaceAfter = 0

    # Para generar el título 
    page_content.append(Paragraph("<b>"+sistema.slogan+"</b>",styleSlogan))

    # Estilos para las sedes y fecha de inicio y fin 
    styleSede= styles['Heading1']
    styleSede.fontName="Times-Roman"
    styleSede.fontSize=12
    styleSede.alignment=TA_CENTER
    styleSede.spaceAfter = 0
    styleSede.spaceBefore = 0
    # styleSede.leading = 20

    page_content.append(Paragraph(sistema.sedes,styleSede))
    # Para agregar fecha
    inicio_dia= datetime.strftime(sistema.fecha_inicio_arbitraje,'%d')
    fin_dia= datetime.strftime(sistema.fecha_fin_arbitraje,'%d')
    mes= datetime.strftime(sistema.fecha_fin_arbitraje,'%m')
    year= datetime.strftime(sistema.fecha_fin_arbitraje,'%Y')
   
    date_event = '%s al %s de %s de %s' % (inicio_dia,fin_dia, MONTH_NAMES[int(mes) - 1], year)
    
    page_content.append(Paragraph(date_event,styleSede))

    # Estilos para la invitación
    styleInvitacion= styles['Heading1']
    styleInvitacion.fontName="Times-Roman"
    styleInvitacion.fontSize=15
    styleInvitacion.alignment=TA_CENTER
    styleInvitacion.spaceAfter = 0

    page_content.append(Paragraph("<b>INVITACIÓN</b>",styleInvitacion))


    # Estilos para el linea de separacion
    styleNombre= styles['Normal']
    styleNombre.fontName="Times-Roman"
    styleNombre.spaceBefore = 0
    styleNombre.spaceAfter = 15
    # Para agregar linea despues del título
    page_content.append(Paragraph("_"*87,styles['Normal']))

    # Estilos para el titulo de la convencion
    # Estilos para el titulo de la convencion
    styleComision= styles['Normal']
    styleComision.fontName="Times-Roman"
    styleComision.alignment=TA_JUSTIFY
    
    # Para dar formato al texto ingresado
    section=""
    if(queryResumen.invitacion):

        for content in queryResumen.invitacion:
            section= section + content
            if (section.find("</p>") >= 0 ): 
                # print section.replace('<br />','')
                section=section.replace('<br />','')
                page_content.append(Paragraph(section, styleComision))
                section=""


    ##############################    Breves palabras    ###################################
    page_content.append(PageBreak())

    sistema= Sistema_asovac.objects.get(pk=arbitraje)
    resumen=Resumen.objects.filter(sistema_id=arbitraje).exists()

    if(resumen == True):
        queryResumen=Resumen.objects.get(sistema_id=arbitraje)
        # archivo_imagen = settings.MEDIA_ROOT+'/'+patrocinadores.url_patrocinadores

    styles=getSampleStyleSheet()
    # start the story...
    page_content.append(NextPageTemplate('remaining_pages'))  #This will load the next PageTemplate with the adjusted Frame. 
    page_content.append(Spacer(1,15))
    # Para agregar cabecera personalizada
    if(queryResumen.cabecera != ""):
        # Estilos para la cabecera
        styleCabecera= styles['BodyText']
        styleCabecera.fontName="Times-Roman"
        styleCabecera.fontSize=12
        styleCabecera.alignment=TA_CENTER
        # styleCabecera.spaceAfter = 0
        # styleCabecera.spaceBefore=0 

        # Para dar formato al texto ingresado
        sectionCabecera=""
        cont=0
        # print queryResumen.cabecera
        if(queryResumen.cabecera):
            for content in queryResumen.cabecera :
                sectionCabecera= sectionCabecera + content
                if (sectionCabecera.find("</p>") >= 0 ): 
                    cont= cont+1
                    # print sectionCabecera.replace('<br />','')
                    # sectionCabecera=sectionCabecera.replace('<br />','')
                    page_content.append(Paragraph("<b>"+sectionCabecera+"</b>", styleCabecera))
                    sectionCabecera="" 
        if(queryResumen.cabecera):
            if cont == 0:
                page_content.append(Paragraph("<b>"+queryResumen.cabecera+"</b>", styleCabecera))
            page_content.append(Spacer(1,10))
    
    # Estilos para la invitación
    styleInvitacion= styles['Heading1']
    styleInvitacion.fontName="Times-Roman"
    styleInvitacion.fontSize=15
    styleInvitacion.alignment=TA_CENTER
    styleInvitacion.spaceAfter = 0

    page_content.append(Paragraph("<b>Palabras</b>",styleInvitacion))
    if(queryResumen.presidente_nombre):
        page_content.append(Paragraph("<b>"+queryResumen.presidente_nombre+"</b>",styleInvitacion))
    
    if(queryResumen.presidente_cargo):
        page_content.append(Paragraph("<b>"+queryResumen.presidente_cargo+"</b>",styleInvitacion))


    # Estilos para el linea de separacion
    styleNombre= styles['Normal']
    styleNombre.fontName="Times-Roman"
    styleNombre.spaceBefore = 0
    styleNombre.spaceAfter = 15
    # Para agregar linea despues del título
    page_content.append(Paragraph("_"*87,styles['Normal']))

    # Estilos para el titulo de la convencion
    # Estilos para el titulo de la convencion
    styleComision= styles['Normal']
    styleComision.fontName="Times-Roman"
    styleComision.alignment=TA_JUSTIFY
    
    # Para dar formato al texto ingresado
    section=""
    if(queryResumen.presidente_contenido):
        for content in queryResumen.presidente_contenido:
            section= section + content
            if (section.find("</p>") >= 0 ): 
                # print section.replace('<br />','')
                section=section.replace('<br />','')
                page_content.append(Paragraph(section, styleComision))
                section="" 
    
    if(queryResumen.secretario_nombre):
        if queryResumen.secretario_nombre != "":
            # Para cargar informacion del secretario
            page_content.append(NextPageTemplate('remaining_pages'))
            page_content.append(PageBreak())
            page_content.append(Spacer(1,15))
            # page_content.append(Spacer(1,15))
            # Para agregar cabecera personalizada
            if(queryResumen.cabecera != ""):
                # Estilos para la cabecera
                styleCabecera= styles['BodyText']
                styleCabecera.fontName="Times-Roman"
                styleCabecera.fontSize=12
                styleCabecera.alignment=TA_CENTER
                # styleCabecera.spaceAfter = 0
                # styleCabecera.spaceBefore=0 

                # Para dar formato al texto ingresado
                sectionCabecera=""
                cont=0
                # print queryResumen.cabecera
                for content in queryResumen.cabecera :
                    sectionCabecera= sectionCabecera + content
                    if (sectionCabecera.find("</p>") >= 0 ): 
                        cont= cont+1
                        # print sectionCabecera.replace('<br />','')
                        # sectionCabecera=sectionCabecera.replace('<br />','')
                        page_content.append(Paragraph("<b>"+sectionCabecera+"</b>", styleCabecera))
                        sectionCabecera="" 
                if cont == 0:
                    page_content.append(Paragraph("<b>"+queryResumen.cabecera+"</b>", styleCabecera))
                page_content.append(Spacer(1,10))

            # Estilos para la invitación
            styleInvitacion= styles['Heading1']
            styleInvitacion.fontName="Times-Roman"
            styleInvitacion.fontSize=15
            styleInvitacion.alignment=TA_CENTER
            styleInvitacion.spaceAfter = 0

            page_content.append(Paragraph("<b>Palabras</b>",styleInvitacion))

            page_content.append(Paragraph("<b>"+queryResumen.secretario_nombre+"</b>",styleInvitacion))
            page_content.append(Paragraph("<b>"+queryResumen.secretario_cargo+"</b>",styleInvitacion))


            # Estilos para el linea de separacion
            styleNombre= styles['Normal']
            styleNombre.fontName="Times-Roman"
            styleNombre.spaceBefore = 0
            styleNombre.spaceAfter = 15
            # Para agregar linea despues del título
            page_content.append(Paragraph("_"*87,styles['Normal']))

            # Estilos para el titulo de la convencion
            # Estilos para el titulo de la convencion
            styleComision= styles['Normal']
            styleComision.fontName="Times-Roman"
            styleComision.alignment=TA_JUSTIFY
            
            # Para dar formato al texto ingresado
            section=""
            for content in queryResumen.presidente_contenido:
                section= section + content
                if (section.find("</p>") >= 0 ): 
                    # print section.replace('<br />','')
                    section=section.replace('<br />','')
                    page_content.append(Paragraph(section, styleComision))
                    section=""


    ##############################         Afiche        ###################################

    if(resumen == True):
        afiche=Resumen.objects.get(sistema_id=arbitraje)
        if(afiche.url_afiche):
            archivo_imagen = settings.MEDIA_ROOT+'/'+afiche.url_afiche

    # frame_first_page =Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id='first')
    
    # add the PageTempaltes to the BaseDocTemplate. You can also modify those to adjust the margin if you need more control over the Frames.
    page_content.append(PageBreak())
    # doc.addPageTemplates([PageTemplate(id='first_page',frames=frame_first_page)])
    styles=getSampleStyleSheet()
    # start the story...

    # page_content.append(NextPageTemplate('first_page'))

    # Estilos para la invitación
    styleAfiche= styles['Heading1']
    styleAfiche.fontName="Times-Roman"
    styleAfiche.fontSize=12
    styleAfiche.alignment=TA_CENTER
    styleAfiche.spaceAfter = 0

    page_content.append(Spacer(1,16))
    if(afiche.afiche_titulo):
        page_content.append(Paragraph(afiche.afiche_titulo,styleAfiche))
        page_content.append(Spacer(1,8))
    
    if(resumen == True):
        img = Image(archivo_imagen,460, 620)
        # img = Image(archivo_imagen,460, 685)
        page_content.append(img)

    ##############################        Eventos        ###################################
    eventos=Evento.objects.filter(sistema_asovac=arbitraje).exists()

    if(eventos == True):
        queryEventos=Evento.objects.filter(sistema_asovac=arbitraje).order_by('fecha_inicio','hora_inicio')

    # Frame(leftMargin,bottomMargin,width,height,leftPadding,rightPadding,topPadding,bottomPadding,id='normal')
    frame_resTrabLib = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id='resTrabLib')
    frame_remaining_pages = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id='remaining')
    # add the PageTempaltes to the BaseDocTemplate. You can also modify those to adjust the margin if you need more control over the Frames.
    doc.addPageTemplates([PageTemplate(id='resTrabLib',frames=frame_resTrabLib, onPage=partial(header_footer,arbitraje=arbitraje))])
    doc.addPageTemplates([PageTemplate(id='remaining_pages',frames=frame_remaining_pages, onPage=partial(header_footer,arbitraje=arbitraje))])
    styles=getSampleStyleSheet()
    # start the story...

    # Para generar título de resumenes de trabajos libres
    page_content.append(NextPageTemplate('resTrabLib'))
    
    styleArea= styles['Heading1']
    styleArea.fontName="Times-Roman"
    styleArea.fontSize=50
    styleArea.alignment=TA_RIGHT
    # styleArea.spaceBefore = 50
    styleArea.textColor = colors.HexColor('#004275')
    page_content.append(Spacer(1,250))
    page_content.append(Paragraph('<b> Programa de </b> <br /> <br /><b> Eventos</b> <br /> <br /><b> Especiales</b>',styleArea))

    page_content.append(PageBreak())

    # Para generar la sección de comisión académica 
    page_content.append(NextPageTemplate('remaining_pages'))  #This will load the next PageTemplate with the adjusted Frame. 
    
    page_content.append(Spacer(1,10))

    # Estilos para el titulo de los eventos
    styleNombre= styles['Title']
    styleNombre.fontName="Times-Roman"
    styleNombre.fontSize=20
    styleNombre.alignment=TA_CENTER
    styleNombre.spaceAfter = 0
    styleNombre.spaceBefore = 0
    styleNombre.textColor = colors.HexColor('#000000')

    page_content.append(Paragraph("<b>Programa General de Eventos Especiales</b>",styleNombre))
    page_content.append(Spacer(1,20))
    # print queryEventos[0].fecha_inicio
    # print queryEventos[0].fecha_inicio.weekday()
    # print DAY_NAMES[queryEventos[0].fecha_inicio.weekday()]
    day_before=''

    # Estilos para el titulo para la fecha de los eventos
    styleTitleTable= styles['BodyText']
    styleTitleTable.fontName="Times-Roman"
    styleTitleTable.fontSize=10
    styleTitleTable.alignment=TA_JUSTIFY
    styleTitleTable.spaceAfter = 0
    styleTitleTable.spaceBefore = 0
    styleTitleTable.textColor = colors.HexColor('#ffffff')

    styleContentTable= styles['Normal']
    styleContentTable.fontName="Times-Roman"
    styleContentTable.fontSize=10
    styleContentTable.alignment=TA_JUSTIFY
    styleContentTable.spaceAfter = 0
    styleContentTable.spaceBefore = 0

    table_title=[]
    table_content=[]
    contItem=0
    has_title=False
    if(eventos == True):
        for event in queryEventos:
            actual_weekday= DAY_NAMES[event.fecha_inicio.weekday()]
            actual_day= str (event.fecha_inicio.day)
            actual_month= MONTH_NAMES[event.fecha_inicio.month-1]
            
            # print event.locacion_evento.lugar
            if (contItem+1) < len(queryEventos):
                day_next=queryEventos[contItem+1].fecha_inicio
                
            else:
                day_next=''
            
            if day_next == event.fecha_inicio:
                # Para agregar el titulo al inicio de la tabla
                if has_title == False:
                    # Para cargar el título de la tabla
                    t1 = Paragraph(('HORARIO').encode('utf-8'), styleTitleTable)
                    t2 = Paragraph('UBICACIÓN'.encode('utf-8'), styleTitleTable)
                    t3 = Paragraph('DESCRIPCIÓN'.encode('utf-8'), styleTitleTable)
                    table_content.append([t1, t2, t3])
                    has_title=True
                
                horario= event.hora_inicio.strftime(" %I:%M%p")+'-'+event.hora_fin.strftime(" %I:%M%p")
                horario= Paragraph((horario).encode('utf-8'), styleContentTable)
                ubicacion=Paragraph((event.locacion_evento.lugar).encode('utf-8'), styleContentTable)
                descripcion=Paragraph((event.descripcion).encode('utf-8'), styleContentTable)

                table_content.append([horario,ubicacion,descripcion])
            else:
                page_content.append(Paragraph(actual_weekday+" "+actual_day+" de "+actual_month,styles['Normal']))
                page_content.append(Spacer(1,5))

                # Para agregar el titulo al inicio de la tabla
                if has_title == False:
                    # Para cargar el título de la tabla
                    t1 = Paragraph(('HORARIO').encode('utf-8'), styleTitleTable)
                    t2 = Paragraph('UBICACIÓN'.encode('utf-8'), styleTitleTable)
                    t3 = Paragraph('DESCRIPCIÓN'.encode('utf-8'), styleTitleTable)
                    table_content.append([t1, t2, t3])

                horario= event.hora_inicio.strftime(" %I:%M%p")+'-'+event.hora_fin.strftime(" %I:%M%p")
                horario=Paragraph((horario).encode('utf-8'), styleContentTable)
                ubicacion=Paragraph((event.locacion_evento.lugar).encode('utf-8'), styleContentTable)
                descripcion=Paragraph((event.descripcion).encode('utf-8'), styleContentTable)

                table_content.append([horario,ubicacion,descripcion])

                tabla = Table(data = table_content,colWidths=(150,150,150),
                            style = [
                                    ('GRID',(0,0),(-1,-1),0.5,colors.HexColor('#cce0ff')),
                                    ('ALIGN', (1,1), (-1,-1), 'RIGHT'),
                                    ('BOX',(0,0),(-1,-1),2,colors.white),
                                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#99c2ff')),
                                    ]
                            )
                
                page_content.append(tabla)
                page_content.append(Spacer(1,10))
                table_content=[]
                has_title= False

            
            # day_before=DAY_NAMES[event.fecha_inicio.weekday()]
            day_before=event.fecha_inicio
            contItem=contItem+1
            # Para agregar contenido de la tabla

    ############################## Resúmenes de trabajos ###################################

    if(resumen == True):
        queryResumen=Resumen.objects.get(sistema_id=arbitraje)
        # archivo_imagen = settings.MEDIA_ROOT+'/'+patrocinadores.url_patrocinadores

    # Frame(leftMargin,bottomMargin,width,height,leftPadding,rightPadding,topPadding,bottomPadding,id='normal')
    frame_resTrabLib = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id='resTrabLib')
    frame_remaining_pages = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id='remaining')
    # add the PageTempaltes to the BaseDocTemplate. You can also modify those to adjust the margin if you need more control over the Frames.
    doc.addPageTemplates([PageTemplate(id='resTrabLib',frames=frame_resTrabLib, onPage=partial(header_footer,arbitraje=arbitraje))])
    doc.addPageTemplates([PageTemplate(id='remaining_pages',frames=frame_remaining_pages, onPage=partial(header_footer,arbitraje=arbitraje))])
    styles=getSampleStyleSheet()
    # start the story...

    # Para generar título de resumenes de trabajos libres
    page_content.append(NextPageTemplate('resTrabLib'))
    
    styleArea= styles['Heading1']
    styleArea.fontName="Times-Roman"
    styleArea.fontSize=50
    styleArea.alignment=TA_RIGHT
    # styleArea.spaceBefore = 50
    styleArea.textColor = colors.HexColor('#004275')
    page_content.append(Spacer(1,250))
    page_content.append(Paragraph('<b> Resúmenes</b> <br /> <br /><b>Trabajos Libres</b>',styleArea))

    page_content.append(PageBreak())

    # Para generar la sección de comisión académica 
    page_content.append(NextPageTemplate('remaining_pages'))  #This will load the next PageTemplate with the adjusted Frame. 
    
    page_content.append(Spacer(1,15))

    # Estilos para el titulo de la convencion
    styleNombre= styles['Title']
    styleNombre.fontName="Times-Roman"
    styleNombre.fontSize=15
    styleNombre.alignment=TA_CENTER
    styleNombre.spaceAfter = 0
    styleNombre.spaceBefore = 0
    styleNombre.textColor = colors.HexColor('#000000')

    page_content.append(Paragraph("<b>COMISIÓN ACADÉMICA</b>",styleNombre))

     # Estilos para los coordinadores de la comision
    styleCoordinadores= styles['Title']
    styleCoordinadores.fontName="Times-Roman"
    styleCoordinadores.fontSize=13
    styleCoordinadores.alignment=TA_CENTER
    styleCoordinadores.spaceAfter = 0
    page_content.append(Paragraph("<b>Coordinadores por Área</b>",styleCoordinadores))

    # Estilos para el titulo de la convencion
    # Estilos para el titulo de la convencion
    styleComision= styles['Normal']
    styleComision.fontName="Times-Roman"
    styleComision.alignment=TA_JUSTIFY
    styleComision.spaceBefore = 10
    styleComision.spaceAfter = 0

    section=""
    if(queryResumen.coordinadores_area):
        for content in queryResumen.coordinadores_area:
            section= section + content
            if (section.find("</p>") >= 0 ): 
                # print section.replace('<br />','')
                # section=section.replace('<br />','')
                page_content.append(Paragraph(section, styleComision))
                section="" 
    
     # Estilos para el titulo de los miembros de la comision
    styleMiembros= styles['Title']
    styleMiembros.fontName="Times-Roman"
    styleMiembros.fontSize=13
    styleMiembros.alignment=TA_CENTER
    styleMiembros.spaceAfter = 0

    if(queryResumen.titulo_comision):
        page_content.append(Paragraph("<b>"+queryResumen.titulo_comision+"</b>",styleMiembros))

    section=""
    if(queryResumen.titulo_comision):
        for content in queryResumen.miembros_comision:
            section= section + content
            if (section.find("</p>") >= 0 ): 
                # print section.replace('<br />','')
                # section=section.replace('<br />','')
                page_content.append(Paragraph(section, styleComision))
                section="" 

    # Para obtener lsitado de areas, subareas y trabajos aceptados
    areas_list = list (Area.objects.all().order_by('nombre'))
    subAreas_list = list (Sub_area.objects.all().order_by('nombre'))
    trabajos_list=getTrabajosSesion(request)
    trabajos_array=[]
    areas_check=[]
    subareas_check=[]
    trabajos_check=[]
    # Para generar arreglo con el resultado de la consulta
    
    for trabajo in trabajos_list:
        trabajos_array.append(trabajo)


    # areas_list.pop("test")
    # Para recorrer el listado de areas, subareas y cargar la información de cada trabajo
    for area in areas_list:
        # print area.nombre
        for subarea in subAreas_list:
            # print subarea.nombre
            for trabajo in trabajos_array:
    
                if (subarea.nombre == trabajo.subarea) and (area.nombre == trabajo.nombre):
                    # Para agregar página con el título del área de los trabajos a cargar
                    if (area.nombre not in areas_check):
                        # print "area sin cargar {}".format(area.nombre)
                        areas_check.append(area.nombre)
                        # Para cargar nueva pagina
                        page_content.append(NextPageTemplate('remaining_pages'))
                        page_content.append(PageBreak())
                        
                        styleArea= styles['Heading1']
                        styleArea.fontName="Times-Roman"
                        styleArea.fontSize=50
                        styleArea.alignment=TA_RIGHT
                        # styleArea.spaceBefore = 50
                        styleArea.textColor = colors.HexColor('#004275')
                        page_content.append(Spacer(1,250))
                        page_content.append(Paragraph('<b> Área</b> <br /> <br /><b>'+area.nombre+'</b>',styleArea))

                    
                    
                    # Para cargar nueva pagina
                    page_content.append(NextPageTemplate('remaining_pages'))
                    page_content.append(PageBreak())
                    # # Para agregar título de cada subarea
                    if (subarea.nombre not in subareas_check):
                        subareas_check.append(subarea.nombre)
                        # Estilos para las subáreas
                        styleSubAreas= styles['Title']
                        styleSubAreas.fontName="Times-Roman"
                        styleSubAreas.fontSize=19
                        styleSubAreas.alignment=TA_CENTER
                        styleSubAreas.spaceAfter = 0
                        styleSubAreas.spaceBefore = 0
                        styleSubAreas.textColor = colors.HexColor('#000000')
                        page_content.append(Spacer(1,15))
                        page_content.append(Paragraph("Sesión <b> "+subarea.nombre+"</b>",styleSubAreas))
                        # Estilos para el linea de separacion
                        styleNombre= styles['Normal']
                        styleNombre.fontName="Times-Roman"
                        # styleNombre.spaceBefore = 0
                        # styleNombre.spaceAfter = 0
                        # Para agregar linea despues del título
                        page_content.append(Paragraph("_"*87,styles['Normal']))
                        
                    page_content.append(Spacer(1,15))

                    # Para cargar contenido de la pagina

                    styleResumen= styles['Title']
                    styleResumen.fontName="Times-Roman"
                    styleResumen.fontSize=9
                    styleResumen.alignment=TA_CENTER
                    styleResumen.spaceAfter = 0
                    styleResumen.spaceBefore = 0

                    page_content.append(Paragraph("<b>("+trabajo.codigo+") "+trabajo.titulo_espanol.upper()+"</b>",styleResumen))
                    if trabajo.titulo_ingles != "":
                        page_content.append(Paragraph("<b>("+trabajo.titulo_ingles+")</b>",styleResumen))
                    page_content.append(Paragraph(trabajo.lista_autores,styleResumen))
                    page_content.append(Paragraph(trabajo.universidad+" "+trabajo.facultad+" "+trabajo.escuela,styleResumen))
                    
                    page_content.append(Spacer(1,15))
                    styleContent= styles['Normal']
                    styleContent.fontName="Times-Roman"
                    styleContent.alignment=TA_JUSTIFY
                    styleContent.spaceBefore = 0
                    styleContent.spaceAfter = 0
                    page_content.append(Paragraph(trabajo.resumen,styleContent))
                    if trabajo.palabras_clave != "":
                        page_content.append(Spacer(1,15))
                        page_content.append(Paragraph("<b>Palabras clave: </b>"+trabajo.palabras_clave,styleContent))

                    trabajos_check.append(trabajo)
                
            
            # eliminar trabajos ya asignados para reducir el recorrido del ciclo for
            for check in trabajos_check:
                trabajos_array.remove(check) 
            trabajos_check=[]



    ############################## Para construir el pdf ###################################
    doc.build(page_content)

    doc = buffer.getvalue()
    buffer.close()
    response.write(doc)
    return response


#---------------------------------------------------------------------------------#
#                Versiones de ejemplo de uso de la libreria                       #
#---------------------------------------------------------------------------------#
    # Version usando BaseDocTempalte
    # response = HttpResponse(content_type='application/pdf')
    # buffer = BytesIO()
    # logo="https://i.imgur.com/G1PXYZU.jpg"
    # img = Image(logo,595, 829.88976378)
    # img = Image(logo,600, 843)
    # # creation of the BaseDocTempalte. showBoundary=0 to hide the debug borders
    # doc = BaseDocTemplate(buffer,showBoundary=1)
    # # create the frames. Here you can adjust the margins
    # # frame_first_page = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id='first')
    # print doc.width
    # print doc.height
    # # Frame(leftMargin,bottomMargin,width,height,leftPadding,rightPadding,topPadding,bottomPadding,id='normal')
    # frame_first_page = Frame(0, 0, 600, 843,0,0,0,0, id='first')
    # frame_remaining_pages = Frame(doc.leftMargin + 1*inch, doc.bottomMargin + 1*inch, doc.width - 2*inch, doc.height - 2*inch, id='remaining')
    # # add the PageTempaltes to the BaseDocTemplate. You can also modify those to adjust the margin if you need more control over the Frames.
    # doc.addPageTemplates([PageTemplate(id='first_page',frames=frame_first_page, onPage=on_first_page),
    #                     PageTemplate(id='remaining_pages',frames=frame_remaining_pages, onPage=on_remaining_pages),
    #                     ])
    # styles=getSampleStyleSheet()
    # # start the story...
    # Elements=[]
    # Elements.append(img)
    # # Elements.apspend(Paragraph("Frame first page!",styles['Normal']))
    # Elements.append(NextPageTemplate('remaining_pages'))  #This will load the next PageTemplate with the adjusted Frame. 
    # Elements.append(PageBreak()) # This will force a page break so you are guarented to get the next PageTemplate/Frame
    # Elements.append(Paragraph("Frame remaining pages!,  "*200,styles['Normal']))
    # #start the construction of the pdf
    # doc.build(Elements)

    # doc = buffer.getvalue()
    # buffer.close()
    # response.write(doc)
    # return response

    # # Version usando SimpleDocTemplate 
    # response = HttpResponse(content_type='application/pdf')
    # buffer = BytesIO()
    # c = SimpleDocTemplate(buffer,
    #                     rightMargin=0,leftMargin=0,
    #                     topMargin=0,bottomMargin=0)
    # sample_style_sheet = getSampleStyleSheet()

    # flowables = []
    # logo="https://i.imgur.com/G1PXYZU.jpg"
    # im = Image(logo,595, 829.88976378)
    # flowables.append(im)

    # paragraph_1 = Paragraph("El título del texto", sample_style_sheet['Heading1'])
    # paragraph_2 = Paragraph(
    #     " texto de prueba para verificar como es el funcionamiento de los parrafos con esta libreria, limita automaticamente el contenido",
    #     sample_style_sheet['BodyText']
    # )
    # flowables.append(paragraph_1)
    # flowables.append(paragraph_2)
    # c.build(flowables,onLaterPages=add_page_number,)
    
    # c = buffer.getvalue()
    # buffer.close()
    # response.write(c)
    # return response

    # # Inicio version con canvas 
    # context = create_common_context(request)
    # resumen=Resumen.objects.filter(sistema_id=arbitraje).exists()
    # sistema=Sistema_asovac.objects.get(id=arbitraje)

    # if(resumen == True):
    #     portada=Resumen.objects.get(sistema_id=arbitraje)
    #     # patrocinadores_imagen = settings.MEDIA_ROOT+'/'+portada.url_patrocinadores
    #     cabecera_imagen = sistema.cabecera
    
    # # Para generar el pdf de la imagen de la portada
    # #Indicamos el tipo de contenido a devolver, en este caso un pdf
    # response = HttpResponse(content_type='application/pdf')
    # # para descarg el pdf
    # # response['Content-Disposition'] = 'attachment; filename="some_file.pdf"'
    # #La clase io.BytesIO permite tratar un array de bytes como un fichero binario, se utiliza como almacenamiento temporal
    # buffer = BytesIO()
    # #Canvas nos permite hacer el reporte con coordenadas X y Y
    # pdf = canvas.Canvas(buffer,pagesize = A4)
    # #Llamo al método cabecera donde están definidos los datos que aparecen en la cabecera del reporte.
    # #Utilizamos el archivo logo_django.png que está guardado en la carpeta media/imagenes
    # #Definimos el tamaño de la imagen a cargar y las coordenadas correspondientes
    # pdf.rect(50, 740, 500, 70)
    # if(resumen == True):
    #     pdf.drawImage(cabecera_imagen, 50,740, 500, 70,preserveAspectRatio=False)
    #     # pdf.drawImage(patrocinadores_imagen, 0,0, 595, 843,preserveAspectRatio=False)
    # #drawString(x,y, texto)
    # pdf.drawString(70,250,"Este es un ejemplo de parrafo en un PDF con la libreria reportlab y python! limite de contenido ")
    #     #Con show page hacemos un corte de página para pasar a la siguiente
    # pdf.showPage()
    # pdf.save()
    # pdf = buffer.getvalue()
    # buffer.close()
    # response.write(pdf)
    # return response